# Databricks notebook source
# MAGIC %md
# MAGIC # Tijuana — Absenteeism Forecast Pipeline
# MAGIC
# MAGIC **Version:** 3.0 — Daily Forecast + Lag-1 Vintage  
# MAGIC **Data Source:** JDBC → `Business_Intelligence.dbo.MX03_HeadcountData_Timestamps`  
# MAGIC **Granularità input:** giornaliera per dipendente → aggregazione giornaliera per turno  
# MAGIC **Freq:** Daily (Business Days) | **Target:** Abs_rate per Turno (A, B, C, D) + General  
# MAGIC **Model:** NeuralProphet (daily, 1 solo modello)  
# MAGIC
# MAGIC **Approccio Vintage (lag-1):**  
# MAGIC > `Forecast_Vintage` = il Forecast che il **run precedente** aveva prodotto per il giorno corrente.  
# MAGIC > Si accumula nella Delta table run dopo run.
# MAGIC
# MAGIC - Query SQL: `Numero→Clerk`, `Turno→Shift`, `Fecha→Date`, `TotalHours=12`, `AbsHours=min(absenteeism,12)` (**esclusi tardes**)  
# MAGIC - Filtro: `Tipo_de_Dia=‘Hábil’`, `year IN (2024,2025,2026)`  
# MAGIC - `yearly_seasonality=8` | `weekly_seasonality=7` | `n_changepoints=5`
# MAGIC - `n_lags=21` (~3 settimane) | `n_forecasts=30` (~30 giorni) | `freq='D'` (plant opera 7/7)
# MAGIC - `quantiles=[0.05, 0.95]` → `Forecast_Lower`/`Forecast_Upper` (PI 90%, v3.1)
# MAGIC - Custom events: **Carnaval_Tijuana, Semana_Santa, Fiestas_Patrias** (date esatte, ±1 giorno)  
# MAGIC - Output: Delta Table `` `sbx-logistics`.kelly.kelly_mx_forecast `` (Power BI)  
# MAGIC - Checkpoint: `/Volumes/sbx-logistics/kelly/kelly_mx_volume/checkpoints/`
# MAGIC
# MAGIC ---
# MAGIC
# MAGIC ### Changelog
# MAGIC
# MAGIC | Versione | Data | Modifica |
# MAGIC |----------|------|----------|
# MAGIC | 3.2 | 2026-07-09 | **Vintage bounds.** Congelati anche `Forecast_Vintage_Lower`/`_Upper` con la stessa logica lag-1 per-shift — misurabile la copertura empirica del PI 90% (schema a 9 colonne). |
# MAGIC | 3.1 | 2026-07-08 | **Schema standard + intervalli di previsione.** `quantiles=[0.05,0.95]` (PI 90%); Delta table con schema standard 7 colonne (`ds, ID, Actual, Forecast_Vintage, Forecast, Forecast_Lower, Forecast_Upper`) — le 5 colonne metadata (Plant, Frequency, Run_Date, Model_Version, Last_Obs_Date) restano solo in `run_history.csv`. Credenziali JDBC e webhook Teams migrati a `dbutils.secrets` (scope `kelly`). Fix: chiamata a `_notify_email` inesistente (NameError sul path stale-data); guard divisione TotalHours=0; vintage read senza `except Exception` generico. Modulo condiviso `common/kelly_common.py`. |
# MAGIC | 3.0 | 2026-05-12 | **Migrazione a frequenza giornaliera (business days).** n_forecasts=30, n_lags=14, weekly_seasonality=True. Nessun gap tra Actual e Forecast in Power BI. |
# MAGIC | 2.3 | 2026-05-12 | Eliminato modello vintage separato. Forecast_Vintage = lag-1 dal run precedente. |
# MAGIC | 2.2 | 2026-05-12 | Production release: Delta table, checkpoint, forecast log, history CSV. |
# MAGIC | 2.1 | 2026-04-29 | Rimossi School_Start_BC/School_End_BC. Ref: `Tijuana_Experiments_v1.0` |
# MAGIC | 2.0 | 2026-04-23 | Migrazione a frequenza settimanale, NeuralProphet multi-step |

# COMMAND ----------

# MAGIC %md
# MAGIC ## 0. Imports

# COMMAND ----------

# DBTITLE 1,Restart Python (pulisce stato corrotto di torch)
dbutils.library.restartPython()

# COMMAND ----------

import os
import gc
import json
import time
import random
import logging
import tempfile
import traceback
from contextlib import contextmanager
from datetime import timedelta, datetime
from pathlib import Path

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["CUDA_VISIBLE_DEVICES"] = ""  # Force CPU-only (avoids NCCL symbol conflict)

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from neuralprophet import NeuralProphet, save as np_save
from warnings import simplefilter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import torch

simplefilter(action="ignore", category=pd.errors.PerformanceWarning)

# Modulo condiviso (repo root)
import sys
_REPO_ROOT = os.path.abspath(os.path.join(os.getcwd(), '..'))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)
from common import kelly_common as kc

print('Libraries loaded.')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Configuration

# COMMAND ----------

from neuralprophet import set_random_seed

# ── Paths (Unity Catalog Volumes) ──
VOLUME_BASE    = "/Volumes/sbx-logistics/kelly/kelly_mx_volume"
OUTPUT_PATH    = Path(f"{VOLUME_BASE}/output")
CHECKPOINT_DIR = Path(f"{VOLUME_BASE}/checkpoints")
LOG_DIR        = Path(f"{VOLUME_BASE}/logs")
FORECAST_LOG_CSV = LOG_DIR / "forecast_log.csv"
HISTORY_CSV      = LOG_DIR / "run_history.csv"

for _d in [OUTPUT_PATH, CHECKPOINT_DIR, LOG_DIR]:
    try:
        _d.mkdir(parents=True, exist_ok=True)
    except OSError:
        print(f"⚠ Impossibile creare {_d} — il volume potrebbe non essere montato.")

# ── JDBC ──
# v3.1: credenziali da secret scope 'kelly' (mai in chiaro nel codice).
# Setup one-off: databricks secrets create-scope kelly
#                databricks secrets put-secret kelly jdbc_user / jdbc_password
JDBC_URL  = "jdbc:sqlserver://10.80.192.78:1433;databaseName=Business_Intelligence"
JDBC_USER = dbutils.secrets.get(scope="kelly", key="jdbc_user")
JDBC_PWD  = dbutils.secrets.get(scope="kelly", key="jdbc_password")

# Query v2.0: solo absenteeism (esclusi tardes), anni 2024-2026
JDBC_QUERY = """
    SELECT
        [Numero]  AS Clerk,
        [Turno]   AS [Shift],
        [Fecha]   AS [Date],
        12        AS TotalHours,
        CASE 
            WHEN [absenteeism] > 12 THEN 12 
            ELSE [absenteeism] 
        END       AS AbsHours
    FROM [dbo].[MX03_HeadcountData_Timestamps]
    WHERE CAST([Tipo_de_Dia] AS VARCHAR(max)) IN ('Hábil', 'habil','h?bil')
      AND [year] IN ('2024', '2025', '2026')
"""

# ── Forecast ──
TARGET_SHIFTS  = ['A', 'B', 'C', 'D']
START_DATE     = pd.Timestamp('2024-01-01')
FORECAST_DAYS  = 30   # orizzonte forecast giornaliero (~30 giorni)
MODEL_VERSION  = 'v3.2'

SEED = 42
set_random_seed(SEED)
print('Configuration OK.')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Load Data
# MAGIC
# MAGIC Fonte: JDBC a `Business_Intelligence.dbo.MX03_HeadcountData_Timestamps` (SQL Server).  
# MAGIC Mapping SQL: `Numero→Clerk, Turno→Shift, Fecha→Date, 12→TotalHours, min(absenteeism,12)→AbsHours`  
# MAGIC **Esclusi `tardes`** — i ritardi non sono assenze vere.  
# MAGIC Filtro: solo `Tipo_de_Dia = 'Hábil'`, anni 2024-2026.  
# MAGIC Dati giornalieri per dipendente → **aggregati a settimana per turno** nella cella successiva.

# COMMAND ----------

df_spark = (
    spark.read.format("jdbc")
    .option("url",      JDBC_URL)
    .option("user",     JDBC_USER)
    .option("password", JDBC_PWD)
    .option("query",    JDBC_QUERY)
    .option("encrypt", "false")
    .option("driver",   "com.microsoft.sqlserver.jdbc.SQLServerDriver")
    .load()
)
df_raw = df_spark.toPandas()

# Tipi
df_raw['Date']       = pd.to_datetime(df_raw['Date'])
df_raw['TotalHours'] = pd.to_numeric(df_raw['TotalHours'], errors='coerce')
df_raw['AbsHours']   = pd.to_numeric(df_raw['AbsHours'],   errors='coerce')

print(f'JDBC rows  : {len(df_raw):,}')
print(f'Periodo    : {df_raw["Date"].min().date()} -> {df_raw["Date"].max().date()}')
print(f'Turni      : {sorted(df_raw["Shift"].unique())}')
print(f'Clerk univ.: {df_raw["Clerk"].nunique()}')

# COMMAND ----------

# DBTITLE 1,Input validation
# ── Validazione input ──────────────────────────────────────────────────────────────
from datetime import datetime

# Microsoft Teams Incoming Webhook URL — v3.1: da secret scope 'kelly'
TEAMS_WEBHOOK_URL = dbutils.secrets.get(scope="kelly", key="teams_webhook_url")

def _notify_teams(title: str, message: str):
    """Invia notifica al canale Teams via Incoming Webhook (helper condiviso)."""
    kc.notify_teams(TEAMS_WEBHOOK_URL, title, message,
                    job="Kelly_MX", notebook="kelly_mx_forecast")

# ── Validazione ──────────────────────────────────────────────────────────────────
_req_cols = {"Clerk", "Shift", "Date", "TotalHours", "AbsHours"}
_miss     = _req_cols - set(df_raw.columns)
if _miss:
    _msg = f"Dati mancanti di colonne obbligatorie: {_miss}"
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)
if len(df_raw) == 0:
    _msg = "La tabella MX03_HeadcountData_Timestamps \u00e8 vuota (0 righe)."
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["TotalHours"].sum() == 0:
    _msg = "Tutti i valori di TotalHours sono 0 \u2014 dati corrotti o export errato."
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["TotalHours"].max() <= 1:
    _msg = (f"TotalHours ha valore massimo {df_raw['TotalHours'].max()} \u2014 "
            f"file probabilmente esportato come flag 0/1 anzich\u00e9 ore reali.")
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["AbsHours"].max() < 0:
    _msg = f"AbsHours ha valori negativi (max={df_raw['AbsHours'].max()}) \u2014 dati corrotti."
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)
if not set(df_raw["Shift"].unique()).intersection(set(TARGET_SHIFTS)):
    _msg = (f"Nessun turno target trovato. Turni nel dataset: {sorted(df_raw['Shift'].unique())}. "
            f"Attesi: {TARGET_SHIFTS}")
    _notify_teams("TIJUANA \u2014 Validazione fallita", _msg)
    raise RuntimeError(_msg)

# v3.1 FIX: il vecchio blocco chiamava _notify_email (funzione inesistente:
# la definizione era commentata) -> NameError garantito sul path stale-data
# che mascherava il RuntimeError diagnostico. Ora helper condiviso.
_max_raw_date = pd.to_datetime(df_raw["Date"]).max()
kc.check_staleness(
    _max_raw_date, max_days=14,
    source_desc="la tabella Business_Intelligence.dbo.MX03_HeadcountData_Timestamps",
    notify=lambda title, msg: _notify_teams(f"TIJUANA \u2014 {title}", msg),
)

print(f"\u2705 Validazione input superata ({len(df_raw):,} righe, {df_raw['Shift'].nunique()} turni)")
# ── Fine validazione ──────────────────────────────────────────────────────────────

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. Clean & Build Dataset (aggregazione giornaliera per turno)

# COMMAND ----------

df_raw['Absenteeism'] = df_raw['AbsHours'] / df_raw['TotalHours'].replace(0, np.nan)

before = len(df_raw)
df = df_raw[
    (df_raw['Shift'].isin(TARGET_SHIFTS)) &
    (df_raw['AbsHours'] >= 0) &
    (df_raw['Absenteeism'] <= 1)
].copy()
print(f'Righe prima del filtro : {before:,}')
print(f'Righe dopo il filtro   : {len(df):,}')
print(f'Righe rimosse          : {before - len(df):,}')
print()

# --- Aggregazione GIORNALIERA per turno (target NeuralProphet) ---
daily = (
    df.groupby(['Date', 'Shift'], as_index=False)
    .agg(
        AbsHours   = ('AbsHours',   'sum'),
        TotalHours = ('TotalHours', 'sum'),
        Headcount  = ('Clerk',      'nunique')
    )
)
# v3.1: guard divisione per zero (coerente con riga 252 e con COL/DA/ATL)
daily['y'] = daily['AbsHours'] / daily['TotalHours'].replace(0, np.nan)
daily = daily.rename(columns={'Date': 'ds', 'Shift': 'ID'})
daily = daily[daily['ds'] >= START_DATE].reset_index(drop=True)

max_date = daily['ds'].max()
min_date = daily['ds'].min()

print(f'Dataset giornaliero: {min_date.date()} -> {max_date.date()}')
print(f'Turni : {sorted(daily["ID"].unique())}')
print(f'Giorni con dati per turno:')
print(daily.groupby('ID')['ds'].count().to_string())
print()
print('Abs_rate medio giornaliero per turno:')
print(daily.groupby('ID')['y'].describe().round(4).to_string())
print()
print(f'Giorni con y=0 per turno:')
for sh in TARGET_SHIFTS:
    data = daily[daily['ID'] == sh]['y']
    print(f'  Shift {sh}: {(data == 0).sum()} zeri su {len(data)} ({(data==0).mean():.1%})')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. Complete Time Series (fill missing business days)

# COMMAND ----------

# Completa serie temporale con TUTTI i giorni (il plant opera 7/7)
full_days = pd.date_range(start=min_date, end=max_date, freq='D')
full_index = pd.MultiIndex.from_product(
    [full_days, daily['ID'].unique()],
    names=['ds', 'ID']
)

df = (
    pd.DataFrame(index=full_index)
    .reset_index()
    .merge(daily[['ds', 'ID', 'y', 'AbsHours', 'TotalHours', 'Headcount']], on=['ds', 'ID'], how='left')
)
df = df[['ds', 'ID', 'y', 'AbsHours', 'TotalHours', 'Headcount']].sort_values(['ds', 'ID']).reset_index(drop=True)

print(f'Dataset completo (daily, 7/7): {df.shape}')
print(f'Range: {df["ds"].min().date()} -> {df["ds"].max().date()}')
print('NaN per turno (giorni senza dati):')
print(df.groupby('ID')['y'].apply(lambda x: x.isna().sum()).to_string())
print()
print('Zero Abs_rate per turno:')
for sh in TARGET_SHIFTS:
    data = df[df['ID'] == sh]['y'].dropna()
    print(f'  Shift {sh}: {(data == 0).sum()} zeri su {len(data)} ({(data==0).mean():.1%})')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. Events — MX Holidays & Custom Events

# COMMAND ----------

# Custom events per Tijuana (v2.1: rimossi School_Start_BC e School_End_BC)
# Validazione multi-stagione ha mostrato che gli eventi School confondono il modello
# Ref: Tijuana_Experiments_v1.0

custom_events = {
    'Carnaval_Tijuana': [
        '2024-02-10', '2025-03-01', '2026-02-14', '2027-02-06',
    ],
    'Semana_Santa': [
        '2024-03-28', '2025-04-17', '2026-04-02', '2027-03-25',
    ],
    'Fiestas_Patrias': [
        '2024-09-16', '2025-09-16', '2026-09-16', '2027-09-16',
    ],
}

# Eventi a livello GIORNALIERO (non più snappati a settimana)
df_events = pd.concat(
    [
        pd.DataFrame({'event': name, 'ds': pd.to_datetime(dates)})
        for name, dates in custom_events.items()
    ],
    ignore_index=True
)

df_events_wide = (
    df_events
    .assign(value=1)
    .pivot_table(index='ds', columns='event', values='value', aggfunc='max')
    .fillna(0)
    .reset_index()
)

df = df.merge(df_events_wide, on='ds', how='left')
event_cols = list(custom_events.keys())
df[event_cols] = df[event_cols].fillna(0).astype(int)

print('Events merged (daily, date esatte).')
print(df[event_cols].sum().to_string())

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6. Train Data Preparation

# COMMAND ----------

# Split: alleniamo su TUTTI i dati fino a max_date
# Rimuoviamo righe con y=NaN PRIMA del fit (approccio ATL)
# NeuralProphet con drop_missing=True su global models genera metrics NaN (bug 0.9.0)
train_df = df[df['y'].notna()].copy()

print(f'max_date : {max_date.date()}')
print(f'train_df : {train_df["ds"].min().date()} -> {train_df["ds"].max().date()}  ({len(train_df):,} righe)')
print(f'Righe NaN rimosse: {len(df) - len(train_df):,}')
print(f'Forecast horizon: {FORECAST_DAYS} giorni')

# COMMAND ----------

# Rimuove colonne non necessarie per NeuralProphet
COLS_DROP = ['AbsHours', 'TotalHours', 'Headcount']
train_df = train_df.drop(columns=COLS_DROP, errors='ignore').reset_index(drop=True)

print('Colonne train_df:', train_df.columns.tolist())

# COMMAND ----------

# MAGIC %md
# MAGIC ## 7. NeuralProphet — Model Definition
# MAGIC
# MAGIC Parametri per forecast **giornaliero** (allineati ad ATL v2.7):  
# MAGIC - `freq = 'D'`: daily (il plant opera 7/7, días hábiles include Sab/Dom)  
# MAGIC - `yearly_seasonality = 8`: Fourier terms (ATL usa 25, ma Tijuana ha serie più corta)  
# MAGIC - `weekly_seasonality = 7`: Fourier terms espliciti per pattern DOW  
# MAGIC - `n_changepoints = 5`: (ATL usa 10, Tijuana ha meno dati)  
# MAGIC - `n_lags = 21`: ~3 settimane di lags giornalieri (identico ad ATL)  
# MAGIC - `n_forecasts = FORECAST_DAYS` (30): ~30 giorni di orizzonte  
# MAGIC - `trend_global_local = local`, `season_global_local = local`  
# MAGIC - **NO `drop_missing`**: righe NaN filtrate a monte (evita bug metrics NaN in NP 0.9.0)  
# MAGIC - Holiday country: MX (±1 giorno)  
# MAGIC - Custom events: Carnaval_Tijuana, Semana_Santa, Fiestas_Patrias

# COMMAND ----------

def build_model():
    """Modello daily per Tijuana — parametri allineati ad ATL v2.7.
    
    Differenze vs ATL:
    - yearly_seasonality=8 (vs 25): serie più corta (~2.3 anni vs ~4 anni)
    - weekly_seasonality=7 (vs 15): sufficiente per pattern DOW
    - n_changepoints=5 (vs 10): meno dati, meno changepoints
    - n_lags=21: identico ad ATL (~3 settimane di memoria AR)
    - NO drop_missing: righe NaN già filtrate in train_df (evita bug metrics NaN)
    """
    m = NeuralProphet(
        n_lags              = 21,    # ~3 settimane di lags giornalieri (come ATL)
        n_forecasts         = FORECAST_DAYS,
        n_changepoints      = 5,
        trend_reg           = 0.5,
        trend_global_local  = 'local',
        seasonality_mode    = 'additive',
        yearly_seasonality  = 8,     # picchi multimodali: Gen, Ago-Ott, Dic
        weekly_seasonality  = 7,     # Fourier terms espliciti per pattern DOW
        daily_seasonality   = False,
        seasonality_reg     = 1,
        season_global_local = 'local',
        # v3.1: intervallo di previsione 90% (pinball loss aggiuntiva sui quantili)
        quantiles           = kc.QUANTILES,
    )
    m = m.add_country_holidays('MX', lower_window=-1, upper_window=1)
    m.add_events(list(custom_events.keys()))
    m.set_plotting_backend('plotly')
    return m

print('Model builder defined (daily, params allineati ad ATL v2.7).')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 8. Fit — Model Training

# COMMAND ----------

import torch, functools

# PyTorch 2.6+: override torch.load per weights_only=False
_orig = torch.serialization.load

@functools.wraps(_orig)
def _patched_load(*args, **kwargs):
    kwargs.setdefault('weights_only', False)
    return _orig(*args, **kwargs)

torch.load = _patched_load

m = build_model()
metrics = m.fit(train_df, freq='D')  # D = daily (plant opera 7/7)

# Validazione training (come ATL)
final_loss = float(metrics['Loss'].iloc[-1]) if 'Loss' in metrics.columns and metrics['Loss'].notna().any() else None
n_epochs   = len(metrics)
print(f'Training completato:')
print(f'   Epochs: {n_epochs}')
if final_loss is not None:
    print(f'   Loss finale: {final_loss:.6f}')
else:
    print(f'   Loss finale: N/A (verifica manuale)')
    # Fallback: conferma parametri learned
    _n_params = sum(p.numel() for p in m.model.parameters() if p.requires_grad)
    print(f'   Parametri trainabili: {_n_params:,}')
print(metrics.tail())

# COMMAND ----------

# DBTITLE 1,Save model checkpoints
# ── Checkpoint: salva modello su UC Volume ──
from neuralprophet import save as np_save

_RUN_TS = datetime.now()
RUN_ID  = _RUN_TS.strftime('%Y%m%d_%H%M%S')

# Garantisci che il volume e la directory esistano
try:
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    spark.sql("CREATE VOLUME IF NOT EXISTS `sbx-logistics`.kelly.kelly_mx_volume")
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)

_ckpt_current = CHECKPOINT_DIR / f'kelly_mx_current_{RUN_ID}.pkl'
np_save(m, str(_ckpt_current))

# Sovrascrivi anche "latest" per accesso rapido
_ckpt_latest = CHECKPOINT_DIR / 'kelly_mx_current_latest.pkl'
np_save(m, str(_ckpt_latest))

print(f'\u2705 Checkpoint salvato:')
print(f'   {_ckpt_current.name}')
print(f'   + latest aggiornato')
print(f'   RUN_ID: {RUN_ID}')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 10. Forecast

# COMMAND ----------

def fill_future_events(future_df):
    cols_to_drop = [c for c in event_cols if c in future_df.columns]
    future_df = future_df.drop(columns=cols_to_drop)
    future_df = future_df.merge(df_events_wide, on='ds', how='left')
    future_df[event_cols] = future_df[event_cols].fillna(0).astype(int)
    return future_df


future = m.make_future_dataframe(train_df, periods=FORECAST_DAYS)
future = fill_future_events(future)

fut_check = future[future['ds'] > max_date][['ds'] + event_cols].drop_duplicates('ds')
n_ev = (fut_check[event_cols].sum(axis=1) > 0).sum()
print(f'Giorni futuri con almeno 1 evento custom: {n_ev}')
if n_ev > 0:
    print(fut_check[fut_check[event_cols].sum(axis=1) > 0][['ds'] + event_cols].to_string(index=False))

forecast = m.predict(future)
print(f'\nForecast generato: {len(forecast):,} righe')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 11. Post-processing

# COMMAND ----------

# v3.1: helper condiviso — per ogni ID get_latest_forecast ('origin-0') + quantili
# 5%/95% -> Forecast_Lower / Forecast_Upper (clip [0,1], no quantile crossing).
df_forecast = kc.extract_latest_forecast(forecast, m, col_name='Forecast')
print(f'Forecast: {df_forecast["ds"].min().date()} -> {df_forecast["ds"].max().date()}  ({len(df_forecast)} righe)')

# COMMAND ----------

# MAGIC %md
# MAGIC ## 12. Final Merge + Vintage Lag-1
# MAGIC
# MAGIC Legge la Delta table del run precedente per recuperare il Forecast che era stato prodotto  
# MAGIC per la settimana corrente. Quel valore diventa `Forecast_Vintage` (lag-1).

# COMMAND ----------

# DBTITLE 1,Final Merge + Vintage Lag-1 (per-shift)
# =============================================================================
# 12. FINAL MERGE + VINTAGE LAG-1 (dalla Delta table del run precedente)
#
# Logica CORRETTA (v3.0.1): per ogni turno, estrai il Forecast dalla tabella
# precedente per le date che ORA hanno un Actual nel run corrente.
# Approccio per-shift: ogni turno ha schedule diverso (A=Mar-Ven, B=Lun-Gio,
# C=Lun/Sab/Dom, D=Ven/Sab/Dom), quindi max_date globale non funziona.
#
# FIX: il vecchio codice usava `prev_df[prev_df['ds'] == max_date]` che
# falliva per C/D perché il Sabato (max_date) era nella LORO zona Actual,
# non Forecast. Ora estraiamo vintage per-shift dalla zona forecast del
# run precedente.
# =============================================================================
TABLE_NAME = '`sbx-logistics`.kelly.kelly_mx_forecast'

# 1. Leggi forecast precedente dalla Delta table.
# v3.1 FIX: None SOLO se la tabella non esiste (primo run); ogni altro errore
# viene rilanciato per non azzerare silenziosamente lo storico Forecast_Vintage.
prev_df = kc.read_delta_or_none(spark, TABLE_NAME)
if prev_df is None or prev_df.empty:
    print('⚠️  Delta table non trovata o vuota (primo run).')
    print('   Forecast_Vintage sarà NaN per questo run.')
    all_vintage = pd.DataFrame(columns=['ds', 'ID'] + kc.VINTAGE_COLS)
else:
    prev_df['ds'] = pd.to_datetime(prev_df['ds'])
    # v3.2: colonne bound mancanti (tabella con schema vecchio) -> NaN
    for _c in ['Forecast_Lower', 'Forecast_Upper'] + kc.VINTAGE_COLS:
        if _c not in prev_df.columns:
            prev_df[_c] = np.nan
    for _c in ['Forecast'] + ['Forecast_Lower', 'Forecast_Upper'] + kc.VINTAGE_COLS:
        prev_df[_c] = pd.to_numeric(prev_df[_c], errors='coerce')

    _LAG1_RENAME = {'Forecast': 'Forecast_Vintage',
                    'Forecast_Lower': 'Forecast_Vintage_Lower',
                    'Forecast_Upper': 'Forecast_Vintage_Upper'}

    # --- NUOVO: Vintage per-shift ---
    # Per ogni turno, prendi le date dove il run precedente aveva un Forecast
    # E che nel run corrente hanno un Actual (= overlap forecast→actual)
    curr_actual_dates = (
        df[df['y'].notna()].groupby('ID')['ds'].apply(set).to_dict()
    )

    new_vintage_rows = []
    for shift_id in TARGET_SHIFTS:
        prev_shift = prev_df[prev_df['ID'] == shift_id]
        # Zona forecast del run precedente per questo turno (point + bound insieme)
        prev_has_forecast = prev_shift[prev_shift['Forecast'].notna()][
            ['ds', 'ID', 'Forecast', 'Forecast_Lower', 'Forecast_Upper']]

        # Date che ORA hanno actual (= possiamo confrontare previsione vs realtà)
        shift_actuals = curr_actual_dates.get(shift_id, set())

        # Overlap: predetto nel run precedente E ora verificabile
        shift_vintage = prev_has_forecast[prev_has_forecast['ds'].isin(shift_actuals)]
        if len(shift_vintage) > 0:
            new_vintage_rows.append(shift_vintage.rename(columns=_LAG1_RENAME))

    if new_vintage_rows:
        vintage_from_prev = pd.concat(new_vintage_rows, ignore_index=True)
    else:
        vintage_from_prev = pd.DataFrame(columns=['ds', 'ID'] + kc.VINTAGE_COLS)

    # Storico vintage accumulato (già consolidato nei run precedenti)
    prev_vintage_history = prev_df[
        (prev_df['Forecast_Vintage'].notna()) &
        (prev_df['ID'].isin(TARGET_SHIFTS))  # solo turni reali, General ricalcolato
    ][['ds', 'ID'] + kc.VINTAGE_COLS]
    
    # Rimuovi duplicati: se un giorno ha già vintage nello storico, non sovrascrivere
    if len(prev_vintage_history) > 0 and len(vintage_from_prev) > 0:
        existing_keys = set(prev_vintage_history[['ds', 'ID']].apply(tuple, axis=1))
        vintage_from_prev = vintage_from_prev[
            ~vintage_from_prev[['ds', 'ID']].apply(tuple, axis=1).isin(existing_keys)
        ]
    
    # Unisci: storico + nuovo vintage
    all_vintage = pd.concat([prev_vintage_history, vintage_from_prev], ignore_index=True)
    
    print(f'\u2705 Delta table letta: {len(prev_df)} righe')
    print(f'   Nuovo vintage (per-shift): {len(vintage_from_prev)} righe')
    print(f'   Storico vintage accumulato: {len(prev_vintage_history)} righe')
    print(f'   Totale vintage: {len(all_vintage)} righe')
    # Dettaglio per turno
    for s in TARGET_SHIFTS:
        n = len(all_vintage[all_vintage['ID'] == s])
        print(f'     {s}: {n} giorni con vintage')

# 2. Costruisci merged_df: Actual + Forecast (nuovo, con bounds) + Forecast_Vintage (lag-1)
merged_df = (
    df[['ds', 'ID', 'y']]
    .merge(df_forecast[['ds', 'ID', 'Forecast', 'Forecast_Lower', 'Forecast_Upper']], on=['ds', 'ID'], how='outer')
    .rename(columns={'y': 'Actual'})
    .sort_values(['ds', 'ID'])
    .reset_index(drop=True)
)

# 3. Aggiungi il trio Forecast_Vintage(_Lower/_Upper) (lag-1 accumulato)
if len(all_vintage) > 0:
    merged_df = merged_df.merge(all_vintage, on=['ds', 'ID'], how='left')
else:
    for _c in kc.VINTAGE_COLS:
        merged_df[_c] = np.nan

# Fix dtype numerico
for _c in kc.NUMERIC_COLS:
    merged_df[_c] = pd.to_numeric(merged_df[_c], errors='coerce')

# 4. General = media dei turni (skipna=True → usa turni disponibili)
# NOTA v3.1: la media dei quantili per-turno NON e' il quantile della media
# (sottostima l'ampiezza vera dell'intervallo di General). Accettabile per
# visualizzazione PBI; da rivedere se il PI di General guida decisioni di staffing.
general = (
    merged_df[merged_df['ID'].isin(TARGET_SHIFTS)]
    .groupby('ds', as_index=False)[kc.NUMERIC_COLS]
    .mean()
    .assign(ID='General')
)

merged_df = (
    pd.concat([merged_df, general], ignore_index=True)
    .sort_values(['ds', 'ID'])
    .reset_index(drop=True)
)

# Verifica vintage per turno
for s in sorted(merged_df['ID'].unique()):
    n_vint = merged_df[(merged_df['ID'] == s) & (merged_df['Forecast_Vintage'].notna())].shape[0]
    if n_vint > 0:
        dates = merged_df[(merged_df['ID'] == s) & (merged_df['Forecast_Vintage'].notna())]['ds']
        print(f'  {s}: {n_vint} vintage days ({dates.min().date()} -> {dates.max().date()})')
    else:
        print(f'  {s}: NESSUN vintage')

merged_df = merged_df[kc.STANDARD_COLS]

print(f'\nmerged_df: {merged_df.shape}')
print(f'ID presenti: {sorted(merged_df["ID"].unique())}')
print(f'Giorni con Forecast_Vintage: {merged_df["Forecast_Vintage"].notna().sum()}')

# Verifica continuità Actual/Forecast
_last_actual = merged_df[merged_df['Actual'].notna()]['ds'].max()
_first_fc = merged_df[(merged_df['Forecast'].notna()) & (merged_df['ds'] > _last_actual)]['ds'].min()
if pd.notna(_first_fc):
    _gap = (_first_fc - _last_actual).days
    if _gap <= 1:
        print(f'\u2705 Continuità OK: ultimo Actual {_last_actual.date()} -> primo Forecast {_first_fc.date()} (gap {_gap}g)')
    else:
        print(f'\u26a0\ufe0f  Gap di {_gap} giorni tra ultimo Actual ({_last_actual.date()}) e primo Forecast ({_first_fc.date()})')

merged_df.tail(10)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 13. Evaluation Metrics

# COMMAND ----------

def calc_metrics(actual: pd.Series, forecast: pd.Series) -> dict:
    """MAE, Bias, RMSE, SMAPE — esclude righe con NaN (helper condiviso)."""
    _m = kc.compute_metrics(actual, forecast)
    return dict(n=_m['N'], MAE=_m['MAE'], Bias=_m['Bias'], RMSE=_m['RMSE'], SMAPE=_m['SMAPE'])


# Valutazione: solo giorni con ENTRAMBI Actual e Forecast_Vintage disponibili
# (lo storico si accumula run dopo run)
eval_df = merged_df[
    merged_df['Actual'].notna() &
    merged_df['Forecast_Vintage'].notna()
].copy()

if len(eval_df) == 0:
    print('\u26a0\ufe0f  Primo run \u2014 nessun Forecast_Vintage disponibile per valutazione.')
    print('   Le metriche saranno disponibili dal secondo run in poi.')
else:
    # Metriche globali
    g = calc_metrics(eval_df['Actual'], eval_df['Forecast_Vintage'])
    print('=== Metriche Forecast Vintage (lag-1 giornaliero) ===')
    print(f'   Giorni valutati: {g["n"]}')
    for k, v in g.items():
        if k != 'n':
            print(f'   {k:>6}: {v}')

    # Metriche per turno
    print('\n=== Per Turno ===')
    rows = []
    for uid in sorted(eval_df['ID'].unique()):
        sub = eval_df[eval_df['ID'] == uid]
        row = calc_metrics(sub['Actual'], sub['Forecast_Vintage'])
        rows.append({'ID': uid, **row})
    id_metrics = pd.DataFrame(rows).set_index('ID')
    print(id_metrics.to_string())

# COMMAND ----------

# MAGIC %md
# MAGIC ## 14. Visualization

# COMMAND ----------

START_PLOT = pd.Timestamp('2025-06-01')
STOP_PLOT  = max_date + pd.Timedelta(days=FORECAST_DAYS)
BG_COLOR   = '#0b1c44'
FORE_ZONE  = '#1a2a5a'

SERIES_STYLE = {
    'Actual'           : dict(color='#76b3fa', linewidth=1.5, alpha=0.90, label='Actual',
                              marker='.', markersize=3, zorder=3),
    'Forecast_Vintage' : dict(color='#00e5a0', linewidth=2.0, alpha=0.95, label='Forecast Vintage (lag-1)',
                              linestyle='--', marker='s', markersize=4, zorder=5),
    'Forecast'         : dict(color='#f7b267', linewidth=1.5, alpha=0.90, label='Forecast',
                              linestyle='-', zorder=4),
}

plot_df = merged_df[merged_df['ds'].between(START_PLOT, STOP_PLOT)].copy()

ids    = sorted(plot_df['ID'].unique())
N_COLS = 2
n_rows = int(np.ceil(len(ids) / N_COLS))

plt.style.use('dark_background')
fig, axes = plt.subplots(n_rows, N_COLS, figsize=(20, n_rows * 4.5))
axes = axes.flatten()

for i, uid in enumerate(ids):
    ax  = axes[i]
    sub = plot_df[plot_df['ID'] == uid]
    ax.set_facecolor(BG_COLOR)

    # Zona forecast
    ax.axvspan(max_date, STOP_PLOT, color=FORE_ZONE, alpha=0.4)

    for col, style in SERIES_STYLE.items():
        s = sub[['ds', col]].dropna(subset=[col])
        if len(s) > 0:
            ax.plot(s['ds'], s[col], **style)

    ax.axvline(max_date, color='white', linewidth=1.0, linestyle=':', alpha=0.5)

    # Label turno
    lbl = 'General' if uid == 'General' else f'Shift {uid}'
    ax.text(0.01, 0.96, lbl, transform=ax.transAxes, color='white', fontsize=11, fontweight='bold',
            va='top', bbox=dict(facecolor='#1a2a5a', alpha=0.8, edgecolor='none', boxstyle='round,pad=0.4'))

    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:.0%}'))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    ax.tick_params(colors='white', labelsize=9)
    ax.grid(color='white', alpha=0.08, linestyle='-')

    if i == 0:
        ax.legend(fontsize=8, loc='upper right', framealpha=0.8,
                  facecolor='#0b1c44', edgecolor='white', labelcolor='white')

for ax in axes[len(ids):]:
    fig.delaxes(ax)

fig.patch.set_facecolor(BG_COLOR)
fig.suptitle('Tijuana \u2014 Absenteeism Forecast Giornaliero per Turno',
             color='white', fontsize=15, fontweight='bold', y=0.995)
plt.tight_layout(rect=[0, 0, 1, 0.97])
plt.show()

# COMMAND ----------

# DBTITLE 1,Delta Table Output section header
# MAGIC %md
# MAGIC ## 16. Output Delta Table (Power BI)
# MAGIC
# MAGIC Scrive `merged_df` su Delta table Unity Catalog per connessione diretta Power BI.
# MAGIC **v3.2 — schema standard 9 colonne** (identico a tutte le geografie):
# MAGIC `ds, ID, Actual, Forecast_Vintage, Forecast, Forecast_Lower, Forecast_Upper, Forecast_Vintage_Lower, Forecast_Vintage_Upper`.
# MAGIC Le colonne metadata restano per-run in `run_history.csv`.

# COMMAND ----------

# DBTITLE 1,Write Delta table for Power BI
# =============================================================================
# 16. DELTA TABLE OUTPUT — Power BI (schema standard 9 colonne, v3.2)
# =============================================================================

# FIX: lasciare NaN per giorni di riposo (días de descanso) anziché forzare a 1.0
# I giorni senza dati nel source SQL (weekend/riposo per turno) restano NaN
# e non vengono mostrati in Power BI — comportamento corretto.
# Non facciamo più fillna(1) — i NaN storici sono giorni non lavorati dal turno

# v3.2 — SCHEMA STANDARD 9 COLONNE (identico a tutte le geografie):
# ds, ID, Actual, Forecast_Vintage, Forecast, Forecast_Lower, Forecast_Upper,
# Forecast_Vintage_Lower, Forecast_Vintage_Upper.
# Le colonne metadata (Plant, Frequency, Run_Date, Model_Version, Last_Obs_Date)
# NON sono piu nella Delta table: restano per-run in run_history.csv.
output_df = kc.finalize_output(merged_df)
_n_out_rows = kc.write_forecast_table(spark, merged_df, TABLE_NAME)

print(f'\u2705 Delta table scritta: {TABLE_NAME}')
print(f'   Righe: {len(output_df):,}')
print(f'   Periodo: {output_df["ds"].min().date()} \u2192 {output_df["ds"].max().date()}')
print(f'   ID: {sorted(output_df["ID"].unique())}')
print(f'   Actual NaN (giorni riposo): {output_df["Actual"].isna().sum()}')
print(f'   Vintage lag-1 disponibili: {output_df["Forecast_Vintage"].notna().sum()}')
print(f'   Run: {_RUN_TS.strftime("%Y-%m-%d")} (metadata in run_history.csv)')

# COMMAND ----------

# DBTITLE 1,Forecast Log section header
# MAGIC %md
# MAGIC ## 17. Forecast Log & Run History
# MAGIC
# MAGIC Archivio congelato delle previsioni (per frozen bias analysis) e audit trail dei run.

# COMMAND ----------

# DBTITLE 1,Forecast log (frozen predictions)
# =============================================================================
# 17. FORECAST LOG — ARCHIVIO CONGELATO DEI FORECAST
# Ogni run appende le nuove previsioni. Quando i dati actual arrivano,
# vengono aggiornati nel log per calcolare il frozen bias.
# =============================================================================
new_log_rows = (
    df_forecast[df_forecast['ds'] > max_date][['ds', 'ID', 'Forecast']]
    .copy()
    .assign(
        run_id        = RUN_ID,
        run_date      = _RUN_TS.strftime('%Y-%m-%d'),
        model_version = MODEL_VERSION,
        horizon_days  = lambda d: (d['ds'] - max_date).dt.days,
        actual_value  = pd.NA,
    )
    .rename(columns={'Forecast': 'forecast_value'})
    [['run_id', 'run_date', 'model_version', 'ds', 'ID',
      'horizon_days', 'forecast_value', 'actual_value']]
)

if FORECAST_LOG_CSV.exists():
    log_df = pd.read_csv(FORECAST_LOG_CSV, parse_dates=['ds'])

    # Aggiorna actual dove disponibili
    actuals_available = (
        merged_df[merged_df['Actual'].notna() & (merged_df['Actual'] < 1)]
        [['ds', 'ID', 'Actual']]
        .set_index(['ds', 'ID'])['Actual']
    )
    idx = pd.MultiIndex.from_arrays([log_df['ds'], log_df['ID']])
    actual_updates = actuals_available.reindex(idx).values
    needs_update   = pd.notna(actual_updates) & log_df['actual_value'].isna()
    log_df.loc[needs_update, 'actual_value'] = actual_updates[needs_update]

    # Deduplica
    existing_keys = set(
        log_df['run_id'] + '|' + log_df['ds'].dt.date.astype(str) + '|' + log_df['ID']
    )
    new_log_rows['_key'] = (
        new_log_rows['run_id'] + '|' +
        new_log_rows['ds'].dt.date.astype(str) + '|' +
        new_log_rows['ID']
    )
    new_log_rows_filtered = new_log_rows[~new_log_rows['_key'].isin(existing_keys)].drop(columns='_key')
    new_log_rows.drop(columns='_key', inplace=True)
    updated_log = pd.concat([log_df, new_log_rows_filtered], ignore_index=True)
else:
    new_log_rows_filtered = new_log_rows.copy()
    updated_log = new_log_rows.copy()

updated_log.to_csv(FORECAST_LOG_CSV, index=False)
n_log_new = len(new_log_rows_filtered)
n_log_tot = len(updated_log)
print(f'\u2705 Forecast log: {FORECAST_LOG_CSV.name} (+{n_log_new} nuove righe, {n_log_tot} totali)')

# COMMAND ----------

# DBTITLE 1,Run history (audit trail)
# =============================================================================
# 18. RUN HISTORY — Una riga per esecuzione, audit trail completo
# =============================================================================

# Metriche (se disponibili)
if len(eval_df) > 0:
    _g = calc_metrics(eval_df['Actual'], eval_df['Forecast_Vintage'])
else:
    _g = dict(n=0, MAE=None, Bias=None, RMSE=None, SMAPE=None)

history_row = pd.DataFrame([{
    'run_id':             RUN_ID,
    'run_datetime':       _RUN_TS.strftime('%Y-%m-%d %H:%M:%S'),
    'model_version':      MODEL_VERSION,
    'last_obs_date':      str(max_date.date()),
    'n_shifts':           len(TARGET_SHIFTS),
    'forecast_days':      FORECAST_DAYS,
    'frequency':          'B (business days)',
    'vintage_type':       'lag-1 (previous run)',
    'vintage_n_days':     _g.get('n'),
    'volatile_mae':       _g.get('MAE'),
    'volatile_bias':      _g.get('Bias'),
    'volatile_rmse':      _g.get('RMSE'),
    'volatile_smape_pct': _g.get('SMAPE'),
    'log_rows_total':     n_log_tot,
    'checkpoint':         str(_ckpt_current),
    'delta_table':        TABLE_NAME,
}])

# Fix FUSE: UC Volumes non supporta mode='a' — read → concat → overwrite
if HISTORY_CSV.exists():
    existing_history = pd.read_csv(HISTORY_CSV)
    history_row = pd.concat([existing_history, history_row], ignore_index=True)

history_row.to_csv(HISTORY_CSV, mode='w', header=True, index=False)

print(f'\u2705 History: {HISTORY_CSV.name} ({len(history_row)} run totali)')
if _g.get('n', 0) > 0:
    print(f'   MAE={_g["MAE"]}  Bias={_g["Bias"]}  SMAPE={_g["SMAPE"]}%  (su {_g["n"]} giorni)')
else:
    print('   (primo run \u2014 metriche non ancora disponibili)')
print(f'\n\U0001f3c1 Pipeline completato con successo.')