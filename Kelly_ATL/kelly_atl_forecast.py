# Databricks notebook source
# MAGIC %md
# MAGIC # Kelly ATL — Databricks Forecast (`kelly_atl_forecast`)
# MAGIC Migrato da Windows Task Scheduler a Databricks notebook. Versione corrente: vedi `MODEL_VERSION` / changelog.
# MAGIC
# MAGIC **Struttura:**
# MAGIC   - Input: SQL Server via JDBC (tabella Operations.dbo.absenteeism_by_dept_area)
# MAGIC   - Struttura ID: Shift{1-4} x Department_Area → raggruppati in Gruppo A (Mar-Ven) e B (Lun/Sab/Dom)
# MAGIC   - Aggregati sintetici: General_A (Shift1+2) e General_B (Shift3+4) — media pesata per roster_hc
# MAGIC   - Modelli separati per gruppo A e B (2 modelli current; vintage = lag-1 dalla Delta table, v2.9)
# MAGIC   - Orizzonte forecast: 30 giorni
# MAGIC   - Holiday country: US
# MAGIC   - Eventi: Easter/Good Friday (calcolati), Eid al-Fitr, Super Bowl, NBA Finals, World Series, March Madness, US Open Tennis, School Start/End
# MAGIC   - Quality flag: semaforo diagnostico (🟢/🟡/🔴) basato su verdetti General_A/B — non bloccante
# MAGIC   - Output: CSV + Delta table (`sbx-logistics`.`kelly-abs-forecast`.kelly_atl_forecast) per Power BI
# MAGIC
# MAGIC ---
# MAGIC ### Changelog
# MAGIC
# MAGIC **v2.9** (2026-07-09) — Vintage lag-1 + vintage bounds
# MAGIC - **Vintage = lag-1 carry-forward dalla Delta table** (come COL/DA/MX/IT, cfr. MX v2.3): eliminati i
# MAGIC   2 modelli vintage trainati — il vintage ora misura cio' che la prod ha realmente pubblicato e il
# MAGIC   job e' ~2x piu veloce. La finestra di valutazione/quality flag si riempie in ~4-5 run settimanali
# MAGIC   dopo lo switch (guard su eval vuota).
# MAGIC - **Vintage bounds**: congelati anche `Forecast_Vintage_Lower`/`_Upper` — misurabile la copertura
# MAGIC   empirica del PI 90% (schema a 9 colonne).
# MAGIC
# MAGIC **v2.8** (2026-07-08) — Schema standard + intervalli di previsione
# MAGIC - **Prediction interval 90%** — `quantiles=[0.05, 0.95]` sui modelli CURRENT (i vintage restano point-only);
# MAGIC   nuove colonne `Forecast_Lower` / `Forecast_Upper` in CSV e Delta table (schema standard 7 colonne).
# MAGIC - **Modulo condiviso `common/kelly_common.py`** — metriche, estrazione forecast+quantili, eventi, notifiche.
# MAGIC - **Fix eventi futuri inerti (bug)** — i 4 `make_future_dataframe` ora ricevono `events_df`:
# MAGIC   prima Easter/Super Bowl/scuola valevano 0 su tutto l'orizzonte di forecast.
# MAGIC - **Fix `fillna(1)` (bug, regressione su v2.6.1)** — il changelog v2.6.1 dichiarava il fillna rimosso ma il
# MAGIC   codice lo faceva ancora in due punti; ora i giorni senza osservazione restano NaN e le righe NaN
# MAGIC   vengono scartate prima del fit (approccio MX). `y>0.65 → NaN` (non piu 1.0).
# MAGIC - **Webhook Teams da secret scope `kelly`** — rimosso URL (con token) hardcoded; rimosso blocco JDBC
# MAGIC   commentato con credenziali in chiaro.
# MAGIC - **Delta write con `overwriteSchema`** — prima mancava: il primo run a 7 colonne sarebbe fallito.
# MAGIC
# MAGIC **v2.7** (2026-05-08) — Semplificazione pipeline, rimozione Champion/Challenger
# MAGIC - **Rimosso Champion/Challenger** — con retraining settimanale e parametri fissi, il C/C è ridondante. Il modello viene sempre deployato se il quality flag non è RED.
# MAGIC - **Production Gate → Quality Flag** — declassato da decisore a semaforo diagnostico (🟢 GREEN / 🟡 YELLOW / 🔴 RED). Non blocca output.
# MAGIC - **GATE_THR → QUALITY_THR** — rinominate le soglie per riflettere il nuovo ruolo.
# MAGIC - **General_A/B pesati** — media pesata per roster_hc (Shift1 Picking 60 HC pesa 3x vs Shift1 Receiving 20 HC).
# MAGIC - **FORECAST_HORIZON = 30** — centralizzato, ridotto da 365 a 30 giorni. Training più veloce (decoder leggero).
# MAGIC - **Eventi migliorati** — Easter/Good Friday calcolati via `dateutil.easter` (zero manutenzione). Ramadan → Eid al-Fitr (singola data). Warning automatico se eventi hardcoded non coprono l'orizzonte.
# MAGIC - **Checkpoint latest only** — sovrascrive ogni run (`_latest.pkl`), niente accumulo file.
# MAGIC - **Delta table output** — forecast scritto in `sbx-logistics.kelly.kelly_atl_forecast` per Power BI.
# MAGIC - **Report Excel eliminato** — ridondante (tutto già in History CSV, Forecast Log, output notebook).
# MAGIC - **Fix NaN metriche Gruppo B** — filtro righe con Forecast_Vintage NaN prima di compute_metrics.
# MAGIC - **Metriche settimanali** — aggiunto breakdown settimana per settimana (globale + per ID).
# MAGIC
# MAGIC **v2.6.2** (2026-04-28) — Parameter optimization + interpolation fix
# MAGIC - **Model A**: `n_lags=21, yearly=25, weekly=15, n_changepoints=10, trend_reg=0.5`
# MAGIC   - Search WMAE: 0.0227 (−28% vs v2.6)
# MAGIC   - CV 5-fold WMAE: 0.0392 ± 0.0120
# MAGIC - **Model B**: `n_lags=21, yearly=True, weekly=30, n_changepoints=5, trend_reg=1.0`
# MAGIC   - Search WMAE: 0.0291 (−40% vs v2.6)
# MAGIC   - CV 5-fold WMAE: 0.0510 ± 0.0115
# MAGIC - **Interpolation fix**: serie giornaliera completa + interpolazione lineare per gap > 30gg
# MAGIC   (chiusure natalizie). NeuralProphet riceve 0 NaN → nessun problema LR finder.
# MAGIC   Valutazione solo su work days reali.
# MAGIC - Analisi: notebook [Kelly ATL - Statistical Analysis and CV]
# MAGIC
# MAGIC **v2.6.1** — Preprocessing fix
# MAGIC - `y > 0.65 → NaN` (non 1.0) per coerenza con post-processing
# MAGIC - `roster_hc=0 → NaN` (no scheduled operations)
# MAGIC - Rimosso `fillna(1)` nel completamento serie — NaN come loss mask
# MAGIC
# MAGIC ---
# MAGIC **TODO — Migliorie pianificate:**
# MAGIC - [x] **Cross-validation rolling** — completata in notebook di analisi (5-fold, holdout 30gg, gap 60gg)
# MAGIC - [x] **Parameter search** — 32 combinazioni grid + parallelo, metrica WMAE
# MAGIC - [x] **Output in Delta table** — `sbx-logistics.kelly.kelly_atl_forecast` per Power BI
# MAGIC - [x] **Uncertainty quantification** — v2.8: `quantiles=[0.05, 0.95]` → `Forecast_Lower`/`Forecast_Upper`
# MAGIC - [ ] **Cutoff Gruppo B** — il start a 2024-01-01 esclude il vecchio regime (2022: 47%, 2023: 36%). Valutare in futuro se estendere/ridurre la finestra man mano che si accumula più storico post-cambio regime
# MAGIC - [x] **Secret scope JDBC** — v2.8: scope `kelly` (jdbc_user / jdbc_password / teams_webhook_url); webhook migrato, JDBC pronto per la riattivazione
# MAGIC - [ ] **MLflow tracking** — loggare parametri, metriche e artefatti in MLflow per confronto esperimenti
# MAGIC - [ ] **Ablation test eventi** — validare quali eventi sportivi impattano realmente l'assenteismo ATL (come fatto per Tijuana)

# COMMAND ----------

# DBTITLE 1,DIAGNOSTIC: torch loading (temp)
dbutils.library.restartPython()

# COMMAND ----------

# DBTITLE 1,Imports
import os
import gc
import sys
import json
import time
import random
import logging
import tempfile
import traceback
from contextlib import contextmanager
from datetime import timedelta, datetime
from pathlib import Path

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

import torch
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from neuralprophet import NeuralProphet, save as np_save
from warnings import simplefilter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

simplefilter(action="ignore", category=pd.errors.PerformanceWarning)

# Modulo condiviso (repo root)
_REPO_ROOT = os.path.abspath(os.path.join(os.getcwd(), ".."))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)
from common import kelly_common as kc

# COMMAND ----------

# DBTITLE 1,Configuration
# MAGIC %md
# MAGIC ## Configuration

# COMMAND ----------

# DBTITLE 1,Configuration: seed, paths, constants, mappings
# =============================================================================
# SEED GLOBALE — garantisce riproducibilità tra run con stessi parametri
# =============================================================================
RANDOM_SEED = 42
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)
torch.manual_seed(RANDOM_SEED)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(RANDOM_SEED)

# =============================================================================
# PERCORSI — Unity Catalog Volumes
# =============================================================================
VOLUME_BASE = kc.volume_base("atl")

OUTPUT_PATH     = Path(f"{VOLUME_BASE}/output")
REPORT_PATH     = Path(f"{VOLUME_BASE}/reports")
PLOT_PATH       = Path(f"{VOLUME_BASE}/plots")
LOG_DIR         = Path(f"{VOLUME_BASE}/logs")
CHECKPOINT_DIR  = Path(f"{VOLUME_BASE}/checkpoints")

HISTORY_CSV          = REPORT_PATH / "kelly_atl_run_history.csv"
FORECAST_LOG_CSV     = REPORT_PATH / "kelly_atl_forecast_log.csv"

# Versione del modello: aggiornare manualmente quando cambiano i config
MODEL_VERSION = "v2.9"

# Orizzonte di forecast (giorni)
FORECAST_HORIZON = 30

# Soglie di qualità — usate per verdetti per ID e quality flag (diagnostica)
QUALITY_THR = {
    "wmae_pp_warn":     4.0,   "wmae_pp_red":      5.0,
    "bias_pp_min":     -2.0,   "bias_pp_ok_min":   0.0,   "bias_pp_max":   2.0,
    "max_err_warn":     5.0,   "max_err_red":      7.0,
    "drift_red":        1.5,
    "under_consec_red": 2,
}

for p in [OUTPUT_PATH, REPORT_PATH, PLOT_PATH, LOG_DIR, CHECKPOINT_DIR]:
    p.mkdir(parents=True, exist_ok=True)

# =============================================================================
# MAPPING DIPARTIMENTI → AREE AGGREGATE
# =============================================================================
DEPT_MAPPING = {
    "CAPACITY TEAM":  "RTS/Cap/Repl",
    "RTS":            "RTS/Cap/Repl",
    "Replen":         "RTS/Cap/Repl",
    "Receiving":      "Receiving/Putaway",
    "Putaway":        "Receiving/Putaway",
    "Picking":        "Picking",
    "Packing":        "Packing",
    "Merge":          "Merge",
    "SIMULATION":     "Sim/Qual/Export/Mil/Sort/Ship",
    "Quality Assurance": "Sim/Qual/Export/Mil/Sort/Ship",
    "Export/Military":"Sim/Qual/Export/Mil/Sort/Ship",
    "SORTATION":      "Sim/Qual/Export/Mil/Sort/Ship",
    "SHIPPING":       "Sim/Qual/Export/Mil/Sort/Ship",
    "NAASC":          "Sim/Qual/Export/Mil/Sort/Ship",
}

# Aree da forecastare (le altre vengono scartate)
AREAS_TO_FORECAST = [
    "Shift1 - Packing",
    "Shift1 - Picking",
    "Shift1 - Receiving/Putaway",
    "Shift2 - Packing",
    "Shift2 - Picking",
    "Shift2 - Receiving/Putaway",
    "Shift3 - Packing",
    "Shift3 - Picking",
    "Shift3 - Receiving/Putaway",
    "Shift4 - Packing",
    "Shift4 - Picking",
]

# Giorni lavorativi per gruppo (dayofweek: 0=Lun, 5=Sab, 6=Dom)
WORK_DAYS = {
    "A": [1, 2, 3, 4],   # Shift1/2 + General_A — Mar-Ven
    "B": [0, 5, 6],      # Shift3/4 + General_B — Lun, Sab, Dom
}

# COMMAND ----------

# DBTITLE 1,Logging and timing helpers
# =============================================================================
# LOGGING — StreamHandler su stdout (Spark Connect non cattura stderr nelle celle)
# =============================================================================
RUN_TS = datetime.now()
RUN_ID = RUN_TS.strftime("%Y%m%d_%H%M%S")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
    force=True,
)
log = logging.getLogger(__name__)

# =============================================================================
# TIMING HELPER
# =============================================================================
_timings: dict[str, float] = {}

@contextmanager
def timed(label: str):
    log.info(f"▶ {label}...")
    t0 = time.perf_counter()
    yield
    elapsed = time.perf_counter() - t0
    _timings[label] = round(elapsed, 2)
    log.info(f"✓ {label} — {elapsed:.1f}s")

script_start = time.perf_counter()

# COMMAND ----------

# DBTITLE 1,Data Loading and Preprocessing
# MAGIC %md
# MAGIC ## 1. Data Loading and Preprocessing

# COMMAND ----------

# DBTITLE 1,JDBC data loading from SQL Server
# =============================================================================
# 1. CARICAMENTO DATI DA CSV (Unity Catalog Volume)
# =============================================================================
# --- JDBC originale (commentato) ---
# v2.8: credenziali rimosse dal sorgente — se si riattiva il JDBC usare:
#     .option("user",     dbutils.secrets.get(scope="kelly", key="jdbc_user"))
#     .option("password", dbutils.secrets.get(scope="kelly", key="jdbc_password"))
# with timed("Caricamento dati da SQL Server"):
#     query = "SELECT * FROM [Operations].[dbo].[absenteeism_by_dept_area]"
#     df_spark = spark.read \
#         .format("jdbc") \
#         .option("url", "jdbc:sqlserver://10.80.192.78:1433;databaseName=Operations") \
#         .option("query", query) \
#         .option("encrypt", "false") \
#         .option("driver", "com.microsoft.sqlserver.jdbc.SQLServerDriver") \
#         .load()
#     df_raw = df_spark.toPandas()
#     df_raw.columns = df_raw.columns.str.strip()
# --- Fine JDBC ---

INPUT_CSV = Path(f"{VOLUME_BASE}/input/Absenteeism.csv")

with timed("Caricamento dati da CSV (Volume)"):
    df_raw = pd.read_csv(INPUT_CSV)
    df_raw.columns = df_raw.columns.str.strip()

log.info(f"Righe caricate da CSV: {len(df_raw)}")
log.info(f"Colonne: {list(df_raw.columns)}")

# COMMAND ----------

# DBTITLE 1,Input validation
# ── Validazione input ──────────────────────────────────────────────────────────────
# Microsoft Teams Incoming Webhook URL — v2.8: da secret scope 'kelly'
# TEAMS_WEBHOOK_URL = dbutils.secrets.get(scope="kelly", key="teams_webhook_url")
TEAMS_WEBHOOK_URL = None  # temporaneamente disabilitato (secret scope da configurare)

def _notify_teams(title: str, message: str, color: str = "FF0000"):
    """Invia notifica al canale Teams via Incoming Webhook (helper condiviso)."""
    if not TEAMS_WEBHOOK_URL:
        return
    kc.notify_teams(TEAMS_WEBHOOK_URL, title, message,
                    job="Kelly_ATL", notebook="kelly_atl_forecast", log=log)

# ── Validazione ──────────────────────────────────────────────────────────────────
_req_cols = {"dt", "roster_hc", "present_hc_with_ot", "shift", "Department_Area"}
_miss     = _req_cols - set(df_raw.columns)
if _miss:
    _msg = f"Dati mancanti di colonne obbligatorie: {_miss}"
    _notify_teams("KELLY ATL — Validazione fallita", _msg)
    raise RuntimeError(_msg)
if len(df_raw) == 0:
    _msg = "La tabella absenteeism_by_dept_area è vuota (0 righe)."
    _notify_teams("KELLY ATL — Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["roster_hc"].sum() == 0:
    _msg = "Tutti i valori di roster_hc sono 0 — dati corrotti o export errato."
    _notify_teams("KELLY ATL — Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["roster_hc"].max() <= 1:
    _msg = (f"roster_hc ha valore massimo {df_raw['roster_hc'].max()} — "
            f"file probabilmente esportato come flag 0/1 anziché headcount reale.")
    _notify_teams("KELLY ATL — Validazione fallita", _msg)
    raise RuntimeError(_msg)
if df_raw["present_hc_with_ot"].max() <= 1:
    _msg = (f"present_hc_with_ot ha valore massimo {df_raw['present_hc_with_ot'].max()} — "
            f"file probabilmente esportato come flag 0/1 anziché headcount reale.")
    _notify_teams("KELLY ATL — Validazione fallita", _msg)
    raise RuntimeError(_msg)
_max_raw_date = pd.to_datetime(df_raw["dt"]).max()
_days_stale = kc.check_staleness(
    _max_raw_date, max_days=14,
    source_desc=f"il file CSV in {VOLUME_BASE}/input/",
    notify=lambda title, msg: _notify_teams(f"KELLY ATL — {title}", msg),
    log=log,
)
# ── Fine validazione ──────────────────────────────────────────────────────────────

# COMMAND ----------

log.info(f"Ultima data disponibile nei dati: {_max_raw_date.date()}")
log.info(f"Giorni di staleness: {_days_stale}")

# COMMAND ----------

# DBTITLE 1,Preprocessing and aggregation
# Rinomina e mapping dipartimenti
df_raw["Department_Area"] = df_raw["Department_Area"].replace(DEPT_MAPPING)
df_raw["shift"] = df_raw["shift"].astype(str)
df_raw.rename(columns={"dt": "ds"}, inplace=True)
df_raw["ds"] = pd.to_datetime(df_raw["ds"])

# Aggregazione per data / area / shift
df_grouped = (
    df_raw
    .groupby(["ds", "Department_Area", "shift"], as_index=False)
    .agg(roster_hc=("roster_hc", "sum"), present_hc_with_ot=("present_hc_with_ot", "sum"))
)

# Calcolo tasso assenteismo: 1 - (presenti / rostered)
# v2.8 FIX (regressione su v2.6.1): y>0.65 → NaN (non 1.0) e roster_hc=0 resta NaN.
# Il vecchio fillna(1) fabbricava giorni al 100% di assenza che entravano nel training.
roster_safe = df_grouped["roster_hc"].replace(0, np.nan)
df_grouped["y"] = 1 - (df_grouped["present_hc_with_ot"] / roster_safe)
df_grouped["y"] = np.where(df_grouped["y"] > 0.65, np.nan, df_grouped["y"])
df_grouped["y"] = df_grouped["y"].clip(lower=0, upper=1)

# Costruzione ID: Shift{n} - {Area}
df_grouped["ID"] = "Shift" + df_grouped["shift"] + " - " + df_grouped["Department_Area"]
df_grouped.drop(columns=["Department_Area", "present_hc_with_ot", "shift"], inplace=True)
df_grouped.sort_values(["ds", "ID"], inplace=True)

# Filtra solo le aree da forecastare
df_grouped = df_grouped[df_grouped["ID"].isin(AREAS_TO_FORECAST)].copy()

n_raw_rows = len(df_raw)
log.info(f"Righe raw: {n_raw_rows} | ID disponibili: {sorted(df_grouped['ID'].unique())}")
log.info(f"roster_hc preservato per media pesata General_A/B")

# COMMAND ----------

# DBTITLE 1,Synthetic Aggregates and Time Series Completion
# MAGIC %md
# MAGIC ## 2-3. Synthetic Aggregates and Time Series Completion

# COMMAND ----------

# DBTITLE 1,General_A/B synthetic aggregates
# =============================================================================
# 2. AGGREGATI SINTETICI: General_A e General_B (media pesata per roster_hc)
# Peso = headcount rostered → aree più grandi impattano proporzionalmente
# Equivale a: 1 - Σ(presenti) / Σ(rostered) a livello di gruppo
# =============================================================================
with timed("Creazione aggregati General_A / General_B (pesati)"):
    def _weighted_general(df_src, prefix, gen_id):
        """Media pesata: sum(y * roster_hc) / sum(roster_hc) per giorno."""
        grp = df_src[df_src["ID"].str.startswith(prefix)].copy()
        grp["_wy"] = grp["y"] * grp["roster_hc"]
        agg = grp.groupby("ds").agg(
            _wy_sum=("_wy", "sum"),
            _w_sum=("roster_hc", "sum")
        )
        agg["y"] = agg["_wy_sum"] / agg["_w_sum"].replace(0, np.nan)
        return agg.reset_index()[["ds", "y"]].assign(ID=gen_id)

    general_A = _weighted_general(df_grouped, ("Shift1", "Shift2"), "General_A")
    general_B = _weighted_general(df_grouped, ("Shift3", "Shift4"), "General_B")

    # Drop roster_hc — non serve più downstream
    df_grouped.drop(columns=["roster_hc"], inplace=True)

    # Aggiungi General_A/B al dataframe
    df_grouped = pd.concat([df_grouped, general_A, general_B], ignore_index=True)
    df_grouped = df_grouped.sort_values(["ds", "ID"]).reset_index(drop=True)

    # Colonna Actual (prima di creare il full index)
    df_grouped.rename(columns={"y": "Actual"}, inplace=True)

log.info(f"ID totali (inclusi General_A/B): {sorted(df_grouped['ID'].unique())}")

# COMMAND ----------

# DBTITLE 1,Time series completion (full date range)
# =============================================================================
# 3. COMPLETAMENTO SERIE TEMPORALE (full date range x tutti gli ID)
# =============================================================================
with timed("Completamento serie temporale"):
    min_date, max_date = df_grouped["ds"].min(), df_grouped["ds"].max()

    full_index = pd.MultiIndex.from_product(
        [pd.date_range(start=min_date, end=max_date, freq="D"), df_grouped["ID"].unique()],
        names=["ds", "ID"],
    )
    # v2.8 FIX: y = Actual senza fillna(1) — i giorni non osservati restano NaN
    # e vengono scartati prima del fit (vedi make_splits).
    df = (
        pd.DataFrame(index=full_index)
        .reset_index()
        .merge(df_grouped[["ds", "ID", "Actual"]], on=["ds", "ID"], how="left")
        .assign(y=lambda x: x["Actual"])
    )
log.info(f"Shape dopo completamento: {df.shape} | Max date: {max_date.date()}")

# COMMAND ----------

# DBTITLE 1,Filtering and Events
# MAGIC %md
# MAGIC ## 4-5. Filtering and Events

# COMMAND ----------

# DBTITLE 1,ID filtering (FIX: protect General_A/B from exclusion)
# =============================================================================
# 4. FILTRAGGIO ID — esclude ID con assenteismo medio > 65% (ultimi 100gg)
#    FIX: General_A e General_B sono protetti dall'esclusione per evitare
#         crash nella Production Gate che assume la loro esistenza.
# =============================================================================
with timed("Filtraggio ID"):
    last_100_start = max_date - pd.Timedelta(days=100)
    ids_to_exclude = (
        df[(df["ds"] >= last_100_start) & (df["ds"] <= max_date) & (df["Actual"] < 1)
           & ~df["ID"].isin(["General_A", "General_B"])]
        .groupby("ID")["Actual"].mean()
        .pipe(lambda s: s[s > 0.65])
        .index
    )
    df = df[~df["ID"].isin(ids_to_exclude)]
    all_IDs = df["ID"].unique()

log.info(f"ID esclusi (Actual medio > 0.65): {list(ids_to_exclude)}")
log.info(f"ID da forecastare ({len(all_IDs)}): {sorted(all_IDs.tolist())}")

# Ricalcola gruppi A/B dopo filtraggio
ids_A = [i for i in all_IDs if i.startswith(("Shift1", "Shift2", "General_A"))]
ids_B = [i for i in all_IDs if i.startswith(("Shift3", "Shift4", "General_B"))]
log.info(f"Gruppo A ({len(ids_A)}): {sorted(ids_A)}")
log.info(f"Gruppo B ({len(ids_B)}): {sorted(ids_B)}")

# COMMAND ----------

# DBTITLE 1,Events preparation
# =============================================================================
# 5. EVENTI
# =============================================================================
from dateutil.easter import easter as _easter_calc

with timed("Preparazione eventi"):
    # ── Easter & Good Friday (calcolati — zero manutenzione) ─────────────────
    _forecast_end_year = (max_date + timedelta(days=FORECAST_HORIZON)).year
    _event_years = list(range(int(min_date.year), _forecast_end_year + 1))

    _easter_dates = [_easter_calc(y) for y in _event_years]
    easter_dates      = [d.strftime("%Y-%m-%d") for d in _easter_dates]
    good_friday_dates = [(d - timedelta(days=2)).strftime("%Y-%m-%d") for d in _easter_dates]

    # ── Eid al-Fitr (fine Ramadan) — non calcolabile, hardcoded ──────────────
    eid_al_fitr_dates = ["2022-05-01", "2023-04-21", "2024-04-09", "2025-03-29", "2026-03-19"]

    # ── Eventi sportivi & scuola — hardcoded, richiedono aggiornamento annuale ─
    _hardcoded_events = {
        "Super Bowl":            ["2022-02-13", "2023-02-12", "2024-02-11", "2025-02-09", "2026-02-15"],
        "NBA Finals Start":      ["2022-06-02", "2023-06-01", "2024-05-30", "2025-06-05", "2026-06-04"],
        "World Series Start":    ["2022-10-28", "2023-10-27", "2024-10-25", "2025-10-24", "2026-10-23"],
        "March Madness Start":   ["2022-03-15", "2023-03-14", "2024-03-19", "2025-03-18", "2026-03-17"],
        "US Open Tennis Finals": ["2022-09-11", "2023-09-10", "2024-09-08", "2025-09-07", "2026-09-13"],
        "School Start":          ["2022-08-01", "2023-08-01", "2024-08-01", "2025-08-04", "2026-08-03"],
        "School End":            ["2022-05-26", "2023-05-26", "2024-05-24", "2025-05-29", "2026-05-22"],
    }

    # ── Assembla dizionario completo ─────────────────────────────────────────
    important_sporting_events = {
        "Easter Sunday": easter_dates,
        "Good Friday":   good_friday_dates,
        "Eid al-Fitr":   eid_al_fitr_dates,
        **_hardcoded_events,
    }

    EVENT_COLS = list(important_sporting_events.keys())

    # ── Warning se eventi hardcoded non coprono l'orizzonte forecast ─────────
    _forecast_horizon_end = max_date + timedelta(days=FORECAST_HORIZON)
    for evt_name, evt_dates in _hardcoded_events.items():
        _last_evt = pd.Timestamp(max(evt_dates))
        if _last_evt < _forecast_horizon_end:
            log.warning(
                f"⚠️  Evento '{evt_name}' copre solo fino a {_last_evt.date()} "
                f"ma il forecast arriva a {_forecast_horizon_end.date()} — aggiornare le date!"
            )

    # ── Pivot wide: una riga per data con colonne binarie per ogni evento ────
    df_events_wide = kc.events_dict_to_wide(important_sporting_events)
    df_events_wide.columns.name = None

    df = df.merge(df_events_wide, on="ds", how="left")
    df[EVENT_COLS] = df[EVENT_COLS].fillna(0)

# COMMAND ----------

# DBTITLE 1,Train/Vintage Splits and Model Definition
# MAGIC %md
# MAGIC ## 6-7. Train/Vintage Splits and Model Definition

# COMMAND ----------

# DBTITLE 1,Train/Vintage splits for groups A and B
# =============================================================================
# 6. SPLIT TRAIN / VINTAGE — per gruppo A e B
# =============================================================================
start_date          = pd.Timestamp("2022-01-01")
split_date          = max_date - timedelta(days=2)
split_date_vintage  = split_date - timedelta(days=FORECAST_HORIZON)


def make_splits(df_full, ids, work_days, start=None):
    """Filtra per IDs, work days e finestra temporale (<= split_date).
    start: data di inizio opzionale; se None usa start_date globale.
    v2.8: righe y=NaN scartate prima del fit (approccio MX) — i giorni non
    osservati non entrano piu nel training come y=1.
    v2.9: ritorna solo il train corrente — il vintage e' lag-1 dalla Delta
    table (niente piu modelli vintage trainati).
    """
    _start = start if start is not None else start_date
    base = df_full[
        df_full["ID"].isin(ids) &
        df_full["ds"].dt.dayofweek.isin(work_days) &
        (df_full["ds"] >= _start) &
        df_full["y"].notna()
    ].drop(columns=["Actual"], errors="ignore").reset_index(drop=True)
    return base[base["ds"] <= split_date]


with timed("Creazione split train A e B"):
    train_df_A = make_splits(df, ids_A, WORK_DAYS["A"])
    # Gruppo B: start 2024-01-01 — esclude vecchio regime (2022: 47%) e transizione (2023: 36%)
    train_df_B = make_splits(df, ids_B, WORK_DAYS["B"], start=pd.Timestamp("2024-01-01"))

log.info(f"Train A: {train_df_A['ds'].min().date()} — {train_df_A['ds'].max().date()} ({len(train_df_A):,} righe, {train_df_A['ID'].nunique()} IDs)")
log.info(f"Train B: {train_df_B['ds'].min().date()} — {train_df_B['ds'].max().date()} ({len(train_df_B):,} righe, {train_df_B['ID'].nunique()} IDs)")

# COMMAND ----------

# DBTITLE 1,Model definition (build_model_A, build_model_B)
# =============================================================================
# 7. DEFINIZIONE MODELLI
# Parametri centralizzati: unica fonte di verità usata sia dai builder
# che dal report Excel (sheet "Config Modelli").
#
# v2.6.2 — Parametri aggiornati da parameter search + 5-fold CV
#           (notebook: "Kelly ATL - Statistical Analysis and CV")
#           Grid: 32 combinazioni × 2 serie, metrica WMAE (2x sotto-stima)
#           CV: 5-fold temporal rolling (holdout 30gg, gap 60gg)
# =============================================================================
_MODEL_A_PARAMS = dict(
    n_lags=21,                  # v2.6: 16 → 21 (1 mese di memoria AR)
    n_forecasts=FORECAST_HORIZON,
    yearly_seasonality=25,      # v2.6: True → 25 Fourier terms
    weekly_seasonality=15,      # v2.6: True → 15 Fourier terms
    n_changepoints=10,          # v2.6: 0 → 10 (regime non statico)
    trend_global_local="local",
    season_global_local="local",
    seasonality_reg=1,
    trend_reg=0.5,              # invariato
)

_MODEL_B_PARAMS = dict(
    n_lags=21,                  # v2.6: 9 → 21
    n_forecasts=FORECAST_HORIZON,
    n_changepoints=5,           # invariato
    yearly_seasonality=True,    # v2.6: 20 → True (auto, 10 terms)
    weekly_seasonality=30,      # invariato
    trend_global_local="local",
    season_global_local="global",
    seasonality_reg=1,
    trend_reg=1.0,              # v2.6: 0.5 → 1.0 (regularizza trend)
    ar_reg=0.5,                 # invariato (non testato in grid)
)


def build_model_A(quantiles: list | None = None) -> NeuralProphet:
    """Modello per Gruppo A (Shift1/2 + General_A — Mar-Ven).

    Parametri ottimizzati via parameter search (WMAE=0.0227, -28% vs v2.6).
    CV 5-fold: WMAE=0.0392 ± 0.0120.
    v2.8: quantiles opzionale (PI 90%). v2.9: esistono solo i modelli current —
    il vintage e' lag-1 dalla Delta table.
    """
    _params = dict(_MODEL_A_PARAMS, **({"quantiles": quantiles} if quantiles else {}))
    m = NeuralProphet(**_params)
    m = m.add_country_holidays("US", lower_window=-1, upper_window=1)
    m.add_events(EVENT_COLS)
    m.set_plotting_backend("plotly")
    return m


def build_model_B(quantiles: list | None = None) -> NeuralProphet:
    """Modello per Gruppo B (Shift3/4 + General_B — Lun/Sab/Dom).

    Parametri ottimizzati via parameter search (WMAE=0.0291, -40% vs v2.6).
    CV 5-fold: WMAE=0.0510 ± 0.0115.
    Serie sparsa (solo 3 giorni/settimana), weekly_seasonality=30 cattura il pattern.
    """
    _params = dict(_MODEL_B_PARAMS, **({"quantiles": quantiles} if quantiles else {}))
    m = NeuralProphet(**_params)
    m = m.add_country_holidays("US", lower_window=-1, upper_window=1)
    m.add_events(EVENT_COLS)
    m.set_plotting_backend("plotly")
    return m

# COMMAND ----------

# DBTITLE 1,Training and Forecast
# MAGIC %md
# MAGIC ## 8-10. Training and Forecast

# COMMAND ----------

# DBTITLE 1,Training current models A and B
# =============================================================================
# 8. TRAINING CURRENT (m_A, m_B)
# =============================================================================
with timed("Training modello current A (Shift1/2 — Mar-Ven)"):
    m_A = build_model_A(quantiles=kc.QUANTILES)
    metrics_A = m_A.fit(train_df_A, freq="D")

final_loss_A = float(metrics_A["Loss"].iloc[-1]) if metrics_A is not None and "Loss" in metrics_A.columns else None
n_epochs_A   = len(metrics_A) if metrics_A is not None else None
log.info(f"Loss finale (current A): {final_loss_A}")

with timed("Training modello current B (Shift3/4 — Lun/Sab/Dom)"):
    m_B = build_model_B(quantiles=kc.QUANTILES)
    metrics_B = m_B.fit(train_df_B, freq="D")

final_loss_B = float(metrics_B["Loss"].iloc[-1]) if metrics_B is not None and "Loss" in metrics_B.columns else None
n_epochs_B   = len(metrics_B) if metrics_B is not None else None
log.info(f"Loss finale (current B): {final_loss_B}")

# Salva checkpoint "latest" (sovrascrive ogni run)
_ckpt_A = CHECKPOINT_DIR / "kelly_atl_model_A_latest.pkl"
_ckpt_B = CHECKPOINT_DIR / "kelly_atl_model_B_latest.pkl"
np_save(m_A, str(_ckpt_A))
np_save(m_B, str(_ckpt_B))
log.info(f"Checkpoint salvati: {_ckpt_A.name} | {_ckpt_B.name}")

# COMMAND ----------

# DBTITLE 1,Vintage = lag-1 (v2.9)
# =============================================================================
# 9. VINTAGE — v2.9: NIENTE PIU MODELLI VINTAGE TRAINATI.
# Forecast_Vintage(_Lower/_Upper) = lag-1 carry-forward dalla Delta table del
# run precedente (stessa semantica di COL/DA/MX/IT — misura cio' che la prod
# ha realmente pubblicato; job ~2x piu veloce). Vedi cella dopo il post-processing.
# =============================================================================

# Libera memoria PyTorch prima dei plot (evita crash C-level del renderer matplotlib)
try:
    torch.cuda.empty_cache()
except Exception:
    pass

# COMMAND ----------

# DBTITLE 1,Forecast generation
# =============================================================================
# 10. FORECAST
# =============================================================================
with timed("Generazione forecast"):
    # v2.8 FIX: events_df con le occorrenze future — senza, gli eventi custom
    # (Easter, Super Bowl, scuola, ...) valevano 0 su tutto l'orizzonte forecast.
    _ev_current = kc.build_future_events_long(
        important_sporting_events, split_date, split_date + timedelta(days=FORECAST_HORIZON))
    _n_ev_c = len(_ev_current) if _ev_current is not None else 0
    log.info(f"Eventi futuri nell'orizzonte: {_n_ev_c}")

    future_A = m_A.make_future_dataframe(train_df_A, periods=FORECAST_HORIZON, events_df=_ev_current)
    future_B = m_B.make_future_dataframe(train_df_B, periods=FORECAST_HORIZON, events_df=_ev_current)

    forecast_A = m_A.predict(future_A)
    forecast_B = m_B.predict(future_B)

    # Concatena per mantenere compatibilità con il post-processing
    forecast = pd.concat([forecast_A, forecast_B], ignore_index=True)

log.info(f"Forecast A: {len(forecast_A):,} righe | Forecast B: {len(forecast_B):,} righe")

# COMMAND ----------

# DBTITLE 1,Post-Processing and Metrics
# MAGIC %md
# MAGIC ## 11-12. Post-Processing and Metrics

# COMMAND ----------

# DBTITLE 1,Post-processing forecast (FIX: soglia 0.65, NaN per off-days)
# =============================================================================
# 11. POST-PROCESSING FORECAST
# Logica off-days: per ogni gruppo, i giorni non lavorativi vengono forzati a NaN
# Soglia 0.65: valori > 0.65 vengono portati a NaN (assenza quasi totale — non forecastabile)
# =============================================================================
off_days_A = [d for d in range(7) if d not in WORK_DAYS["A"]]  # [0, 5, 6]
off_days_B = [d for d in range(7) if d not in WORK_DAYS["B"]]  # [1, 2, 3, 4]


def _group_off_mask(frame: pd.DataFrame) -> pd.Series:
    """True nei giorni off per il gruppo dell'ID (non operativi = nessun forecast)."""
    dow = frame["ds"].dt.dayofweek
    mask_A_off = dow.isin(off_days_A) & frame["ID"].str.startswith(("Shift1", "Shift2", "General_A"))
    mask_B_off = dow.isin(off_days_B) & frame["ID"].str.startswith(("Shift3", "Shift4", "General_B"))
    return mask_A_off | mask_B_off


def _postprocess(forecast_df: pd.DataFrame, model_A: NeuralProphet, model_B: NeuralProphet,
                 col_name: str, with_bounds: bool = False) -> pd.DataFrame:
    """
    Per ogni uid: get_latest_forecast con il modello corretto (A o B) via helper
    condiviso (v2.8: estrae anche i quantili se with_bounds), clip [0,1],
    soglia 0.65 → NaN, NaN nei giorni off per gruppo (non operativi = nessun forecast).
    """
    _model_for_id = lambda uid: model_A if uid.startswith(("Shift1", "Shift2", "General_A")) else model_B
    out = kc.extract_latest_forecast(forecast_df, _model_for_id, col_name, with_bounds=with_bounds)

    out[col_name] = np.where(out[col_name] > 0.65, np.nan, out[col_name])
    out.loc[_group_off_mask(out), col_name] = np.nan

    if with_bounds:
        out = kc.mask_bounds_like_point(out, col_name, f"{col_name}_Lower", f"{col_name}_Upper")
    return out


with timed("Post-processing forecast + vintage lag-1"):
    df_forecast = _postprocess(forecast, m_A, m_B, col_name="Forecast", with_bounds=True)

    # Costruzione merged_df: Actual + Forecast (+ bounds)
    df_actual = df[["ds", "ID", "Actual"]].copy()

    merged_df = (
        df_actual
        .merge(df_forecast[["ds", "ID", "Forecast", "Forecast_Lower", "Forecast_Upper"]], on=["ds", "ID"], how="outer")
    )

    # -- v2.9: Vintage = lag-1 carry-forward dalla Delta table -----------------
    # Congela Forecast(_Lower/_Upper) del run precedente per le date trascorse
    # nel trio Forecast_Vintage*, mantenendo il vintage gia' accumulato.
    FREEZE_UNTIL = pd.Timestamp.today().normalize()
    prev_df = kc.read_delta_or_none(spark, kc.forecast_table("atl"))
    vintage_all, _vmeta = kc.carry_forward_vintage(prev_df, FREEZE_UNTIL)

    merged_df = (
        merged_df
        .merge(vintage_all, on=["ds", "ID"], how="left")
        [kc.STANDARD_COLS]
        .sort_values(["ds", "ID"]).reset_index(drop=True)
    )

    # Ri-applica le maschere off-day per gruppo anche al vintage
    _off = _group_off_mask(merged_df)
    for _c in kc.VINTAGE_COLS:
        merged_df.loc[_off, _c] = np.nan
    merged_df = kc.mask_bounds_like_point(
        merged_df, "Forecast_Vintage", "Forecast_Vintage_Lower", "Forecast_Vintage_Upper")

    _lvd = _vmeta["last_vintage_date"]
    log.info(f"Vintage lag-1: last_vintage_date={_lvd.date() if pd.notna(_lvd) else 'N/A'} "
             f"| punti congelati={_vmeta['n_frozen']} "
             f"| vintage totale={merged_df['Forecast_Vintage'].notna().sum()}")

# COMMAND ----------

# DBTITLE 1,Metrics evaluation
# =============================================================================
# 12. METRICHE DI VALUTAZIONE — GLOBALI
# Valutazione su finestra vintage (split_date_vintage → split_date),
# solo giorni lavorativi per ciascun gruppo.
# =============================================================================
with timed("Calcolo metriche"):
    eval_raw = merged_df[
        (merged_df["ds"] > split_date_vintage) &
        (merged_df["ds"] <= split_date)
    ].copy()

    eval_raw["dow"] = eval_raw["ds"].dt.dayofweek
    df_eval = eval_raw[
        (eval_raw["ID"].str.startswith(("Shift1", "Shift2", "General_A")) & eval_raw["dow"].isin(WORK_DAYS["A"])) |
        (eval_raw["ID"].str.startswith(("Shift3", "Shift4", "General_B")) & eval_raw["dow"].isin(WORK_DAYS["B"]))
    ].copy()
    # Escludi giorni con Actual=1 (non lavorativi rimasti)
    df_eval = df_eval[df_eval["Actual"] < 1]

    # Filtra righe con Forecast_Vintage NaN (coda non coperta dal forecast horizon)
    _n_before = len(df_eval)
    df_eval = df_eval[df_eval["Forecast_Vintage"].notna()].copy()
    _n_dropped = _n_before - len(df_eval)
    if _n_dropped > 0:
        log.warning(
            f"⚠️  {_n_dropped} righe escluse dalla valutazione (Forecast_Vintage NaN — "
            f"coda eval window oltre copertura forecast {FORECAST_HORIZON}gg)"
        )

    # v2.8: helper condiviso (stessa formula, NaN-pair-safe)
    compute_metrics = kc.compute_metrics

    # v2.9: dopo lo switch a vintage lag-1 la finestra eval si riempie in
    # ~4-5 run settimanali — con eval vuota le metriche vengono saltate.
    _METRIC_KEYS = ["MAE", "Bias", "RMSE", "SMAPE", "WMAE", "N"]
    if df_eval.empty:
        log.warning("⚠️  Nessun punto valutabile (vintage lag-1 in accumulo) — metriche saltate per questo run")
        global_metrics = compute_metrics(pd.Series(dtype=float), pd.Series(dtype=float))
        per_id_metrics = pd.DataFrame(columns=["ID"] + _METRIC_KEYS + ["Actual_Mean_Last30d"])
        weekly_metrics = pd.DataFrame(columns=["year_week"] + _METRIC_KEYS)
        weekly_by_id   = pd.DataFrame(columns=["year_week", "ID"] + _METRIC_KEYS)
    else:
        global_metrics = compute_metrics(df_eval["Actual"], df_eval["Forecast_Vintage"])

        per_id_metrics = (
            df_eval.groupby("ID")
            .apply(lambda g: pd.Series(compute_metrics(g["Actual"], g["Forecast_Vintage"])))
            .reset_index()
        )

        last_30 = (
            merged_df[
                (merged_df["ds"] > split_date_vintage) &
                (merged_df["ds"] <= split_date) &
                (merged_df["Actual"] < 1)
            ]
            .groupby("ID")["Actual"].mean().round(4).rename("Actual_Mean_Last30d")
        )
        per_id_metrics = per_id_metrics.merge(last_30.reset_index(), on="ID", how="left")

        # Metriche settimanali
        _iso = df_eval["ds"].dt.isocalendar()
        df_eval["year_week"] = _iso["year"].astype(str) + "-W" + _iso["week"].astype(str).str.zfill(2)

        weekly_metrics = (
            df_eval.groupby("year_week")
            .apply(lambda g: pd.Series(compute_metrics(g["Actual"], g["Forecast_Vintage"])))
            .reset_index()
            .sort_values("year_week")
        )

        weekly_by_id = (
            df_eval.groupby(["year_week", "ID"])
            .apply(lambda g: pd.Series(compute_metrics(g["Actual"], g["Forecast_Vintage"])))
            .reset_index()
            .sort_values(["year_week", "ID"])
        )

# ── Display metriche globali ────────────────────────────────────────────────────
print("=" * 70)
print("METRICHE GLOBALI (holdout: {} → {})".format(
    split_date_vintage.date(), split_date.date()))
print("=" * 70)
print(f"  MAE:    {global_metrics['MAE']:.4f}    ({global_metrics['MAE']*100:.2f} pp)")
print(f"  Bias:   {global_metrics['Bias']:+.4f}   ({global_metrics['Bias']*100:+.2f} pp)")
print(f"  RMSE:   {global_metrics['RMSE']:.4f}")
print(f"  WMAE:   {global_metrics['WMAE']:.4f}    ({global_metrics['WMAE']*100:.2f} pp)")
print(f"  SMAPE:  {global_metrics['SMAPE']:.2f}%")
print(f"  N obs:  {global_metrics['N']}")
if _n_dropped > 0:
    print(f"  ⚠️  Escluse: {_n_dropped} righe (NaN coverage)")
print()

# ── Tabella metriche per ID ──────────────────────────────────────────────────
print("METRICHE PER ID:")
print("-" * 70)
display(per_id_metrics.sort_values("WMAE", ascending=False))
print()

# ── Valutazione settimana per settimana (globale) ──────────────────────────
print("METRICHE SETTIMANALI (globale):")
print("-" * 70)
display(weekly_metrics)
print()

# ── Valutazione settimana per settimana (per ID) ───────────────────────────
print("METRICHE SETTIMANALI PER ID:")
print("-" * 70)
display(weekly_by_id)

# COMMAND ----------

# DBTITLE 1,Verdicts, Production Gate, Champion/Challenger
# MAGIC %md
# MAGIC ## 13-14. Verdicts and Quality Flag

# COMMAND ----------

# DBTITLE 1,Verdicts per ID (FIX: explicit bias > max case)
# =============================================================================
# 13. VERDETTI PER ID — Goldstar + Guardrails (diagnostica)
# =============================================================================
with timed("Verdetti per ID"):
    eval_verd = df_eval.copy()
    _iso = eval_verd["ds"].dt.isocalendar()
    eval_verd["year_week"] = _iso["year"].astype(str) + "-W" + _iso["week"].astype(str).str.zfill(2)

    def _s3(val, warn, red):
        return "RED" if val >= red else ("WARN" if val >= warn else "OK")

    per_id_verdicts: dict = {}

    for _uid in sorted(eval_verd["ID"].unique()):
        _sub = eval_verd[eval_verd["ID"] == _uid].copy()
        _sub["error_pp"]     = (_sub["Forecast_Vintage"] - _sub["Actual"]) * 100
        _sub["abs_error_pp"] = _sub["error_pp"].abs()
        _sub["under"]        = _sub["Forecast_Vintage"] < _sub["Actual"]

        _mae_pp  = float(_sub["abs_error_pp"].mean())
        _bias_pp = float(_sub["error_pp"].mean())
        _rmse_pp = float(np.sqrt((_sub["error_pp"] ** 2).mean()))
        _w       = np.where(_sub["under"], 2, 1)
        _wmae_pp = float((_w * _sub["abs_error_pp"]).sum() / _w.sum())

        _weekly = (
            _sub.groupby("year_week")
            .agg(Actual=("Actual", "mean"), Forecast=("Forecast_Vintage", "mean"))
            .sort_index().reset_index()
        )
        _weekly["error_pp"]   = (_weekly["Forecast"] - _weekly["Actual"]) * 100
        _weekly["abs_err_pp"] = _weekly["error_pp"].abs()
        _weekly["under"]      = _weekly["Forecast"] < _weekly["Actual"]

        _n_under = int(_weekly["under"].sum())
        _n_weeks = len(_weekly)
        _max_consec = _cur = 0
        for _u in _weekly["under"]:
            _cur = _cur + 1 if _u else 0
            _max_consec = max(_max_consec, _cur)

        _max_err_pp = float(_weekly["abs_err_pp"].max()) if _n_weeks > 0 else float("nan")

        if _n_weeks >= 4:
            _first2   = float(_weekly.iloc[:2]["abs_err_pp"].mean())
            _last2    = float(_weekly.iloc[-2:]["abs_err_pp"].mean())
            _drift_pp = _last2 - _first2
        else:
            _drift_pp = float("nan")

        _wmae_s  = _s3(_wmae_pp, QUALITY_THR["wmae_pp_warn"], QUALITY_THR["wmae_pp_red"])
        _bias_s  = ("OK"   if QUALITY_THR["bias_pp_ok_min"] <= _bias_pp <= QUALITY_THR["bias_pp_max"]
                    else "WARN" if QUALITY_THR["bias_pp_min"]    <= _bias_pp <  QUALITY_THR["bias_pp_ok_min"]
                    else "RED")  # covers both < bias_pp_min AND > bias_pp_max
        _under_s = ("OK"  if _n_under <= 1
                    else "RED"  if _max_consec >= QUALITY_THR["under_consec_red"]
                    else "WARN")
        _err_s   = (_s3(_max_err_pp, QUALITY_THR["max_err_warn"], QUALITY_THR["max_err_red"])
                    if not np.isnan(_max_err_pp) else "OK")
        _drift_s = ("RED" if not np.isnan(_drift_pp) and _drift_pp > QUALITY_THR["drift_red"]
                    else "OK")

        if _bias_s == "RED":
            _verdict = "REVISE_MODEL_Bias"
        elif any(_s == "RED" for _s in [_wmae_s, _under_s, _err_s, _drift_s]):
            _failed  = [_n for _n, _s in [("WMAE", _wmae_s), ("Under", _under_s),
                                           ("MaxErr", _err_s), ("Drift", _drift_s)] if _s == "RED"]
            _verdict = f"REVISE_MODEL_{'_'.join(_failed)}"
        elif any(_s == "WARN" for _s in [_wmae_s, _bias_s, _under_s, _err_s]):
            _verdict = "MONITOR"
        else:
            _verdict = "PRODUCTION_READY"

        per_id_verdicts[_uid] = {
            "verdict":     _verdict,
            "wmae_pp":     round(_wmae_pp, 4),
            "bias_pp":     round(_bias_pp, 4),
            "mae_pp":      round(_mae_pp, 4),
            "rmse_pp":     round(_rmse_pp, 4),
            "max_err_pp":  round(_max_err_pp, 4) if not np.isnan(_max_err_pp) else None,
            "drift_pp":    round(_drift_pp, 4)   if not np.isnan(_drift_pp)   else None,
            "n_under":     _n_under, "n_weeks": _n_weeks, "max_consec": _max_consec,
            "wmae_s":      _wmae_s,  "bias_s":  _bias_s,  "under_s":    _under_s,
            "err_s":       _err_s,   "drift_s": _drift_s,
        }

    n_ids_revise  = sum(1 for v in per_id_verdicts.values() if v["verdict"].startswith("REVISE"))
    n_ids_monitor = sum(1 for v in per_id_verdicts.values() if v["verdict"] == "MONITOR")
    n_ids_ready   = sum(1 for v in per_id_verdicts.values() if v["verdict"] == "PRODUCTION_READY")
    log.info(f"Verdetti per ID — REVISE: {n_ids_revise}  MONITOR: {n_ids_monitor}  READY: {n_ids_ready}")

    # Print
    print("VERDETTO PER ID:")
    print("-" * 70)
    if per_id_verdicts:
        display(pd.DataFrame(per_id_verdicts).T.sort_values("verdict").reset_index().rename(columns={"index": "ID"}))
    else:
        print("(nessun verdetto — vintage lag-1 in accumulo)")

# COMMAND ----------

# DBTITLE 1,Production gate
# =============================================================================
# 14. QUALITY FLAG — semaforo diagnostico (non bloccante)
# Riassume lo stato qualitativo basato sui verdetti di General_A/B.
# =============================================================================
with timed("Quality flag"):
    # Bias globale calcolato su General_A e General_B (aggregati pesati)
    _gate_eval = df_eval[df_eval["ID"].isin(["General_A", "General_B"])]
    if len(_gate_eval) > 0:
        _global_bias_pp = float((_gate_eval["Forecast_Vintage"] - _gate_eval["Actual"]).mean()) * 100
    else:
        log.warning("⚠️  Nessuna osservazione per General_A/B nel holdout — bias fallback a globale")
        _global_bias_pp = global_metrics["Bias"] * 100

    # Semaforo basato sui verdetti di General_A e General_B
    _verd_A = per_id_verdicts.get("General_A", {}).get("verdict", "PRODUCTION_READY")
    _verd_B = per_id_verdicts.get("General_B", {}).get("verdict", "PRODUCTION_READY")

    if _verd_A.startswith("REVISE") or _verd_B.startswith("REVISE"):
        quality_flag = "🔴 RED"
        flag_reason  = f"General_A={_verd_A}, General_B={_verd_B}, bias={_global_bias_pp:+.2f}pp"
    elif _verd_A == "MONITOR" or _verd_B == "MONITOR":
        quality_flag = "🟡 YELLOW"
        flag_reason  = f"General_A={_verd_A}, General_B={_verd_B}, bias={_global_bias_pp:+.2f}pp"
    else:
        quality_flag = "🟢 GREEN"
        flag_reason  = f"General_A e General_B OK, bias={_global_bias_pp:+.2f}pp"

    log.info(f"Quality flag: {quality_flag} — {flag_reason}")
    print(f"\n{'='*70}")
    print(f"QUALITY FLAG: {quality_flag}")
    print(f"  {flag_reason}")
    print(f"{'='*70}")

# COMMAND ----------

# DBTITLE 1,Output: Excel, Forecast Log, Frozen Bias, Report, History
# MAGIC %md
# MAGIC ## 16-20. Output: Excel, Forecast Log, Frozen Bias, Report, History

# COMMAND ----------

# DBTITLE 1,Save Excel data (section 16)
# =============================================================================
# 16. SALVATAGGIO EXCEL DATI
# =============================================================================
with timed("Salvataggio Excel dati"):
    file_date    = max_date.strftime("%m-%d-%Y")
    output_file  = OUTPUT_PATH / f"Kelly_ATL_v26_DB_{file_date}.xlsx"
    # fillna(1) solo per le date storiche (≤ max_date): off-days senza osservazione reale.
    # Dopo max_date Actual rimane NaN — non esistono dati reali e non vanno simulati.
    # (Convenzione di sola visualizzazione output — non tocca il training.)
    _actual_out = merged_df["Actual"].copy()
    _hist_mask  = merged_df["ds"] <= max_date
    _actual_out[_hist_mask] = _actual_out[_hist_mask].fillna(1)
    _export_df = kc.finalize_output(merged_df.assign(Actual=_actual_out))
    _export_df.to_csv(output_file.with_suffix('.csv'), index=False)

log.info(f"Dati salvati: {output_file}")

# Scrivi anche su Delta table per Power BI — v2.9: schema standard 9 colonne
# (kc.STANDARD_COLS: point + bounds + vintage trio), round(4),
# overwrite + overwriteSchema.
_n_rows = kc.write_forecast_table(spark, merged_df.assign(Actual=_actual_out),
                                  kc.forecast_table("atl"))
log.info(f"Delta table kelly_atl_forecast scritta: {_n_rows} righe")

# COMMAND ----------

# DBTITLE 1,Forecast log (section 17)
# =============================================================================
# 17. FORECAST LOG — ARCHIVIO CONGELATO DEI FORECAST
# =============================================================================
with timed("Aggiornamento forecast log"):
    new_log_rows = (
        df_forecast[df_forecast["ds"] > split_date][["ds", "ID", "Forecast"]]
        .copy()
        .assign(
            run_id=RUN_ID,
            run_date=RUN_TS.strftime("%Y-%m-%d"),
            model_version=MODEL_VERSION,
            horizon_days=lambda d: (d["ds"] - pd.Timestamp(split_date.date())).dt.days,
            actual_value=pd.NA,
        )
        .rename(columns={"Forecast": "forecast_value"})
        [["run_id", "run_date", "model_version", "ds", "ID", "horizon_days", "forecast_value", "actual_value"]]
    )

    if FORECAST_LOG_CSV.exists():
        log_df = pd.read_csv(FORECAST_LOG_CSV, parse_dates=["ds"])

        actuals_available = (
            merged_df[merged_df["Actual"] < 1][["ds", "ID", "Actual"]]
            .set_index(["ds", "ID"])["Actual"]
        )
        idx = pd.MultiIndex.from_arrays([log_df["ds"], log_df["ID"]])
        actual_updates = actuals_available.reindex(idx).values
        needs_update   = pd.notna(actual_updates) & log_df["actual_value"].isna()
        log_df.loc[needs_update, "actual_value"] = actual_updates[needs_update]

        existing_keys = set(
            log_df["run_id"] + "|" + log_df["ds"].dt.date.astype(str) + "|" + log_df["ID"]
        )
        new_log_rows["_key"] = (
            new_log_rows["run_id"] + "|" +
            new_log_rows["ds"].dt.date.astype(str) + "|" +
            new_log_rows["ID"]
        )
        new_log_rows_filtered = new_log_rows[
            ~new_log_rows["_key"].isin(existing_keys)
        ].drop(columns="_key")
        new_log_rows.drop(columns="_key", inplace=True)
        updated_log = pd.concat([log_df, new_log_rows_filtered], ignore_index=True)
    else:
        new_log_rows_filtered = new_log_rows.copy()
        updated_log = new_log_rows.copy()

    updated_log.to_csv(FORECAST_LOG_CSV, index=False)
    n_log_new_rows = len(new_log_rows_filtered)
    n_log_total    = len(updated_log)
    log.info(f"Forecast log aggiornato: {FORECAST_LOG_CSV} (+{n_log_new_rows} nuove righe, {n_log_total} totali)")

# COMMAND ----------

# DBTITLE 1,Frozen bias (section 18)
# =============================================================================
# 18. FROZEN BIAS
# =============================================================================
with timed("Calcolo frozen bias"):
    HORIZON_BANDS = {
        "h01-07":  (1,    7),
        "h08-14":  (8,   14),
        "h15-28":  (15,  28),
        "h29-56":  (29,  56),
        "h57+":    (57, 9999),
        "all":     (1,  9999),
    }

    frozen_metrics_by_horizon: dict[str, dict] = {}
    n_log_rows_with_actuals = 0

    if FORECAST_LOG_CSV.exists():
        log_eval = pd.read_csv(FORECAST_LOG_CSV, parse_dates=["ds"])
        log_eval = log_eval[
            log_eval["actual_value"].notna() &
            (log_eval["actual_value"].astype(float) < 1) &
            log_eval["forecast_value"].notna()
        ].copy()
        log_eval["actual_value"]   = log_eval["actual_value"].astype(float)
        log_eval["forecast_value"] = log_eval["forecast_value"].astype(float)
        n_log_rows_with_actuals    = len(log_eval)

        if n_log_rows_with_actuals > 0:
            for band, (lo, hi) in HORIZON_BANDS.items():
                band_df = log_eval[log_eval["horizon_days"].between(lo, hi)]
                if len(band_df) >= 5:
                    frozen_metrics_by_horizon[band] = compute_metrics(
                        band_df["actual_value"], band_df["forecast_value"]
                    )

    if frozen_metrics_by_horizon:
        log.info("Frozen Bias per orizzonte:")
        for band, fm in frozen_metrics_by_horizon.items():
            log.info(f"  {band}: Bias={fm['Bias']:+.4f}  MAE={fm['MAE']:.4f}  SMAPE={fm['SMAPE']}%  N={fm['N']}")
    else:
        log.info("Frozen Bias: non ancora disponibile (log insufficiente — almeno 2 run necessari)")

# COMMAND ----------

# DBTITLE 1,History CSV and checkpoints index (section 20)
# =============================================================================
# 20. HISTORY CSV — una riga per run, audit trail completo
# =============================================================================
_fb_all = frozen_metrics_by_horizon.get("all", {})
_fb_h1  = frozen_metrics_by_horizon.get("h01-07", {})
_fb_h4  = frozen_metrics_by_horizon.get("h29-56", {})

total_elapsed = round(time.perf_counter() - script_start, 1)

history_row = pd.DataFrame([{
    # Identificazione run
    "run_id":               RUN_ID,
    "run_datetime":         RUN_TS.strftime("%Y-%m-%d %H:%M:%S"),
    "model_version":        MODEL_VERSION,
    "last_obs_date":        str(max_date.date()),
    "n_ids":                len(all_IDs),
    "ids_excluded":         ", ".join(ids_to_exclude),
    # Finestra di valutazione
    "eval_from":            str(split_date_vintage.date()),
    "eval_to":              str(split_date.date()),
    # Metriche volatile (holdout corrente)
    "volatile_mae":         global_metrics["MAE"],
    "volatile_bias":        global_metrics["Bias"],
    "volatile_rmse":        global_metrics["RMSE"],
    "volatile_wmae":        global_metrics["WMAE"],
    "volatile_smape_pct":   global_metrics["SMAPE"],
    "volatile_n_eval_obs":  global_metrics["N"],
    # Frozen bias (out-of-sample reale)
    "frozen_bias_all":      _fb_all.get("Bias"),
    "frozen_mae_all":       _fb_all.get("MAE"),
    "frozen_wmae_all":      _fb_all.get("WMAE"),
    "frozen_smape_all":     _fb_all.get("SMAPE"),
    "frozen_n_all":         _fb_all.get("N"),
    "frozen_bias_h1w":      _fb_h1.get("Bias"),
    "frozen_mae_h1w":       _fb_h1.get("MAE"),
    "frozen_bias_h4w":      _fb_h4.get("Bias"),
    "frozen_mae_h4w":       _fb_h4.get("MAE"),
    "log_rows_with_actuals": n_log_rows_with_actuals,
    # Quality flag
    "quality_flag":         quality_flag,
    "flag_reason":          flag_reason,
    "bias_pp":              round(_global_bias_pp, 4),
    "n_ids_revise":         n_ids_revise,
    "n_ids_monitor":        n_ids_monitor,
    "n_ids_ready":          n_ids_ready,
    # Training
    "loss_current_A":       final_loss_A,
    "loss_current_B":       final_loss_B,
    # v2.9: niente piu modelli vintage trainati (vintage = lag-1 dalla Delta table)
    "vintage_frozen_points": _vmeta["n_frozen"],
    "checkpoint_A":         str(_ckpt_A),
    "checkpoint_B":         str(_ckpt_B),
    # Timing
    "time_training_current_A_sec": _timings.get("Training modello current A (Shift1/2 \u2014 Mar-Ven)"),
    "time_training_current_B_sec": _timings.get("Training modello current B (Shift3/4 \u2014 Lun/Sab/Dom)"),
    "time_total_sec":       total_elapsed,
    # Output
    "output_file":          str(output_file),
}])

# Fix FUSE: UC Volumes non supporta mode='a' — read → concat → overwrite
if HISTORY_CSV.exists():
    existing_history = pd.read_csv(HISTORY_CSV)
    history_row = pd.concat([existing_history, history_row], ignore_index=True)

history_row.to_csv(HISTORY_CSV, mode="w", header=True, index=False)

log.info(f"History aggiornata: {HISTORY_CSV}")
log.info(f"Script completato in {total_elapsed}s.")
